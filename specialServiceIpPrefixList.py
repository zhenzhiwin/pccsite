import openpyxl
import os


def getServiceListByList(sheet, startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    NameList = []
    # firstLineServiceId = sheet.cell(row=startRow, column=serviceId_col).value
    retList = []
    layerLag = "L7"
    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value
        NameList.append(serviceName)
        retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    NameList = list(set(NameList))
    retList = list(set(retList))

    return retList, NameList


def getTheServiceIpPrefixListDict(service_ip_prefix_list_dict, name_list, configList):
    for name in name_list:
        service_ip_list = []
        for i in range(0, len(configList)):
            if 'ip-prefix-list "IPL_' + name in configList[i] and "create" in configList[i]:
                temp_list = []
                temp_list.append(configList[i].replace("\n", ""))
                k = i
                for j in range(k + 1, len(configList)):
                    if "exit" in configList[j]:
                        break
                    if "name" in configList[j]:
                        temp_list.append(configList[j].replace("\n", ""))
                service_ip_list.append(temp_list)
        service_ip_prefix_list_dict[name] = service_ip_list


def serviceTop(service_name_list):
    topList = ["mgspqwdx_00", "txsp_00", "ks_00", "blbl_00", "mgtv_00", "wyyyy_00", "pptv_00", "zsyyt_01",
               "pushmail_01", "fx_01", "mgsp_00", "mgzb_00", "mgyy_00", "mgyd_00"]
    retList = []
    for i in range(0, len(topList)):
        for serviceList in service_name_list:
            if serviceList[0][3] == topList[i]:
                tmplist = serviceList
                retList.append(tmplist)
                service_name_list.remove(tmplist)
    retList = retList + service_name_list
    return retList


def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []

    newRetList = serviceTop(retList)
    return newRetList


def arrangeTheList_2(rtList):
    new_ret_list = []
    for lst in rtList:
        tmplist = []
        for tup in lst:
            if tup[1] == "新增":
                tmplist.append(tup)
        for tup in lst:
            if tup[1] == "删除":
                tmplist.append(tup)
        new_ret_list.append(tmplist)
    return new_ret_list


def ip_prefix_list_is_full(cfg_ipPrefixList):
    global ip_prefix_list_max_number
    for lst in cfg_ipPrefixList:
        if len(lst) < ip_prefix_list_max_number + 1:
            return False
    return True


def addIpPrefixListCommand(clst, sName, ipStr):
    if "/" not in ipStr:
        ipStr = ipStr + "/32"
    clst.append("prefix " + ipStr + ' name "' + sName + '"' + "\n")


def createTheCaixinIpPrefixList(serviceName, serviceListAdd, command_List, service_IpPrefixList):
    postFixNum = -1

    if len(service_IpPrefixList) == 0:
        tl = []
        postFixNum = 1
        if postFixNum < 10:
            ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_0' + str(postFixNum) + '"'
        else:
            ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_' + str(postFixNum) + '"'
        tl.append(ipliststr)
        service_IpPrefixList.append(tl)

        while len(serviceListAdd) != 0:
            for lst in service_IpPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = serviceListAdd[0]
                    lst.append(serviceListAdd[0])
                    serviceListAdd.remove(serviceListAdd[0])
                    if len(serviceListAdd) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(service_IpPrefixList) == True:
                tl = []
                postFixNum += 1
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_' + str(postFixNum) + '"'
                tl.append(ipliststr)
                service_IpPrefixList.append(tl)

    else:
        # 循环直到用户需要添加iplist(addList)列表中的数据添加完才跳出循环
        while len(serviceListAdd) != 0:
            for lst in service_IpPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = serviceListAdd[0]
                    lst.append(serviceListAdd[0])
                    serviceListAdd.remove(serviceListAdd[0])
                    if len(serviceListAdd) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(service_IpPrefixList) == True:
                tl = []
                postFixNum = len(service_IpPrefixList) + 1
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "IPL_' + serviceName + '_' + str(postFixNum) + '"'
                tl.append(ipliststr)
                service_IpPrefixList.append(tl)


def getTheServiceAddIpPrefixListDict(service_add_IpPrefixList_dict, result_list):
    for lst in result_list:
        # print(lst[0][3],lst[0])
        tmplist = []
        for tup in lst:
            tmplist.append(tup[4])
        service_add_IpPrefixList_dict[lst[0][3]] = tmplist


def getTheCompatibleEntryIdByDict():
    global serviceEntryIdDict
    global allEntryIdList
    global allEntryIdDict
    serviceCaseStr = "no_head"
    retId = -1

    for i in range(20501, 60000):
        if i not in allEntryIdDict[serviceCaseStr]:
            retId = i
            allEntryIdDict[serviceCaseStr].append(retId)
            break
    return retId


def createEntry(ip_list_name, service_name, command_list):
    entryId = getTheCompatibleEntryIdByDict()
    command_list.append("exit all\n")
    command_list.append("configure application-assurance group 1:1 policy\n")
    command_list.append("app-filter\n")
    command_list.append("entry " + str(entryId) + " create\n")
    command_list.append('server-address eq ' + ip_list_name + "\n")
    command_list.append('application "APP_' + service_name + '"\n')
    command_list.append("no shutdown\n")
    command_list.append("\n")


def gen_spec(configList, path, ):
    global ip_prefix_list_max_number
    ip_prefix_list_max_number = 50
    global allEntryIdList
    global serviceDict
    serviceDict = {}
    global serviceEntryIdDict
    serviceEntryIdDict = {}
    global allEntryIdDict
    allEntryIdDict = {}
    global servicePortListDict
    servicePortListDict = {}
    global portListCommandList
    portListCommandList = []

    ccl7_cfg = open(path + '\\configureL7.log', 'r')
    ccl7_cfg_list = ccl7_cfg.readlines()
    ccl7_cfg.close()
    allEntryIdList = eval(ccl7_cfg_list[0])
    serviceDict = eval(ccl7_cfg_list[1])
    serviceEntryIdDict = eval(ccl7_cfg_list[2])
    allEntryIdDict = eval(ccl7_cfg_list[3])
    servicePortListDict = eval(ccl7_cfg_list[4])
    commandList = []
    if os.path.exists(path + "\\内容计费整理L7_specialService.xlsx"):
        excel_path = path + "\\内容计费整理L7_specialService.xlsx"
        tab = "L7"
    elif os.path.exists(path + "\\内容计费整理L34_specialService.xlsx"):
        excel_path = path + "\\内容计费整理L34_specialService.xlsx"
        tab = "L34"
    # chargingContextLog_path = "E:\processL347\问题配置\XIMSAEGW0CBNK-config.txt"
    # configFile = open(chargingContextLog_path, 'r')
    # configList = configFile.readlines()
    # configFile.close()
    #print("excel_path",excel_path,tab)
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel[tab]

    serviceList, nameList = getServiceListByList(sheet, 1)
    resultList = arrangeTheList(serviceList)
    resultList = arrangeTheList_2(resultList)

    serviceAddIpPrefixListDict = {}
    getTheServiceAddIpPrefixListDict(serviceAddIpPrefixListDict, resultList)
    # print(serviceAddIpPrefixListDict)

    serviceIpPrefixListDict = {}
    # print("name list is ",nameList)
    getTheServiceIpPrefixListDict(serviceIpPrefixListDict, nameList, configList)
    # print("处理前")
    # for key in serviceIpPrefixListDict:
    #     print(key)
    #     for lst in serviceIpPrefixListDict[key]:
    #         for line in lst:
    #             print(line)

    for key in serviceAddIpPrefixListDict:
        createTheCaixinIpPrefixList(key, serviceAddIpPrefixListDict[key], commandList, serviceIpPrefixListDict[key])
    # print("处理后")
    for key in serviceIpPrefixListDict:
        # print(key)
        for lst in serviceIpPrefixListDict[key]:
            for line in lst:
                # print(line)
                if "create" in line:
                    commandList.append('exit all\n')
                    commandList.append('configure application-assurance group 1:1\n')
                    commandList.append(line.replace(" create", "").replace("  ", "") + "\n\n")
                elif "ip-prefix-list" in line and "create" not in line:
                    createEntry(line, key, commandList)
                    commandList.append('exit all\n')
                    commandList.append('configure application-assurance group 1:1\n')
                    commandList.append(line.replace(" ", "") + "\n\n")
                elif "name" not in line:
                    addIpPrefixListCommand(commandList, key, line)

    text_cfg = []
    text_cfg.append(str(allEntryIdList) + "\n")
    text_cfg.append(str(serviceDict) + "\n")
    text_cfg.append(str(serviceEntryIdDict) + "\n")
    text_cfg.append(str(allEntryIdDict) + "\n")
    text_cfg.append(str(servicePortListDict) + "\n")

    file = open(path + "\\configureL7.log", "w")
    file.writelines(text_cfg)
    file.close()

    fo = open(path + "\\脚本文件\\特殊业务的ip_prefix_list(例如腾讯视频L3地址放ipprefixlist里).txt", "w")
    fo.writelines(commandList)
    fo.close()
