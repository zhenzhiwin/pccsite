import openpyxl
import ChargingContextAai


def getServiceListByList(sheet, startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if portNumberL4 != None:
            if portsplit_2 in str(portNumberL4):
                portstr = portNumberL4.split(portsplit_2)[0]
                portL4List.append(portNumberL4.split(portsplit_2)[1])
            else:
                pass
            if portsplit_1 in str(portstr):
                portstr = portstr.split(portsplit_1)
                for p in portstr:
                    portL4List.append(p)
            else:
                portL4List.append(portstr)
            for p in portL4List:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
        else:
            retList.append(
                (layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
        if layerLag == "L7" and serviceName == None:
            pass
    retList = list(set(retList))

    return retList


def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList


def arrangeTheList_2(rtList):
    new_ret_list = []
    for lst in rtList:
        tmplist = []
        for tup in lst:
            if tup[1] == "新增":
                tmplist.append(tup)
        for tup in lst:
            if tup[1] == "删除":
                tmplist.append(tup)
        new_ret_list.append(tmplist)
    return new_ret_list


def getTheAllServiceUrl(service_name):
    global configList
    retList = []

    return retList


def addServiceUrlIpListToDict(lst):
    global serviceUrlIpListDict
    urlIpListDict = {}

    serviceName = lst[0][3]

    urlList = []
    for tup in lst:
        urlList.append(tup[7])
    urlList = list(set(urlList))
    for url in urlList:
        urlIpListDict[url] = []
    serviceUrlIpListDict[serviceName] = urlIpListDict
    for urlKey in serviceUrlIpListDict[serviceName]:
        ipPrefixList = []
        ipPrefixList = getIpPrefixListByUrl_serviceName(urlKey, serviceName)
        urlIpListDict[urlKey] = ipPrefixList


def getIpPrefixListByUrl_serviceName(url, service_name):
    global configList
    retList = []
    http_host, http_url, port = ChargingContextAai.processUrl(url)
    http_host = 'expression 1 http-host eq "' + http_host + '"'

    if http_url != None:
        http_url = 'expression 2 http-uri eq "' + http_url + '"'
    for i in range(0, len(configList)):
        if http_host in configList[i]:
            if http_url != None and http_url in configList[i + 1]:
                k = i
                tmplist = []
                for j in range(k, len(configList)):
                    if 'no shutdown' in configList[j]:
                        break
                    if 'server-address eq ip-prefix-list "app_' + service_name in configList[j] and "_HeaderEnrich" in \
                            configList[j]:
                        ipListStr = configList[j].split("server-address eq ")[1].replace("\n", "")
                        tmplist = getTheIpPrefixList(ipListStr)
                if len(tmplist) != 0:
                    retList.append(tmplist)
            elif http_url == None:
                k = i
                tmplist = []
                for j in range(k, len(configList)):
                    if 'no shutdown' in configList[j]:
                        break
                    if 'server-address eq ip-prefix-list "app_' + service_name in configList[j] and "_HeaderEnrich" in \
                            configList[j]:
                        ipListStr = configList[j].split("server-address eq ")[1].replace("\n", "")
                        tmplist = getTheIpPrefixList(ipListStr)
                if len(tmplist) != 0:
                    retList.append(tmplist)
    return retList


def getTheIpPrefixList(ip_list_str):
    global configList
    retList = []
    retList.append(ip_list_str + " create")
    for i in range(0, len(configList)):
        if ip_list_str + ' create' in configList[i]:
            f = i
            for j in range(f, len(configList)):
                if "exit" in configList[j]:
                    e = j
                    break
    for i in range(f, e):
        if 'prefix ' in configList[i]:
            #ipstr = configList[i].replace("\n", "").split("prefix ")[1].split(" name")[0].replace("/32", "")
            ipstr = configList[i].replace("\n", "").replace("    ", "")
            retList.append(ipstr)
    return retList


def setServicePostFixNum(service_name, postList):
    global configList
    for text in configList:
        if 'ip-prefix-list "app_' + service_name + '_' in text and "create" in text and "_HeaderEnrich" in text:
            postfixnumstr = text.split('ip-prefix-list "app_' + service_name + '_')[1].split('" create')[0]
            postList.append(int(postfixnumstr))


def addServiceUrlIpListToUserDict(lst):
    global serviceUrlIpListUserDict
    global del_serviceUrlIpListUserDict
    global servce_ipostfix_num
    serviceName = lst[0][3]

    if serviceName not in servce_ipostfix_num:
        servce_ipostfix_num[serviceName] = []
        setServicePostFixNum(serviceName, servce_ipostfix_num[serviceName])

    urlIpListDict = {}
    urlList = []
    for tup in lst:
        urlList.append(tup[7])
    urlList = list(set(urlList))
    for url in urlList:
        urlIpListDict[url] = []

    for url in urlList:
        tmplst = []
        delTmplst = []
        for tup in lst:
            if tup[7] == url and tup[1] == "新增":
                tmplst.append(tup[4])
            # if tup[7] == url and tup[1]=="删除":
            #   delTmplst.append(tup[4])
        urlIpListDict[url] = tmplst
        # del_urlIpListDict[url] = delTmplst
    serviceUrlIpListUserDict[serviceName] = urlIpListDict
    # del_serviceUrlIpListUserDict[serviceName] = del_urlIpListDict



def addCommandTocommandList(comlst, serverName, url, addList):
    global serviceUrlIpListDict
    global ip_prefix_list_max_number
    global servce_ipostfix_num

    config_ipPrefixList = serviceUrlIpListDict[serverName][url]

    # 若配置列表为空则表示配置文件中无该业务的ip_prefix_list

    if len(config_ipPrefixList) == 0:
        tl = []
        if len(servce_ipostfix_num[serverName]) == 0:
            postFixNum = 1
            servce_ipostfix_num[serverName].append(postFixNum)
        else:
            postFixNum = max(servce_ipostfix_num[serverName]) + 1
            servce_ipostfix_num[serverName].append(postFixNum)
        if postFixNum < 10:
            ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '_HeaderEnrich"'
        else:
            ipliststr = 'ip-prefix-list "app_' + serverName + '_' + str(postFixNum) + '_HeaderEnrich"'
        tl.append(ipliststr)
        #entryId = getTheCompatibleEntryIdByDict()

        #createIpPrefixListEntry(comlst, serverName, url, ipliststr, entryId)
        config_ipPrefixList.append(tl)
        while len(addList) != 0:
            for lst in config_ipPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = addList[0]
                    lst.append(addList[0])
                    addList.remove(addList[0])
                    #addIpPrefixListCommand(comlst, serverName, lst[0], addIpStr)
                    if len(addList) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(config_ipPrefixList) == True:
                tl = []
                if len(servce_ipostfix_num[serverName]) != 0:
                    postFixNum = max(servce_ipostfix_num[serverName]) + 1
                    servce_ipostfix_num[serverName].append(postFixNum)
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '_HeaderEnrich"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_' + str(postFixNum) + '_HeaderEnrich"'
                tl.append(ipliststr)
                #entryId = getTheCompatibleEntryIdByDict()
                #createIpPrefixListEntry(comlst, serverName, url, ipliststr, entryId)
                config_ipPrefixList.append(tl)
    else:
        # 循环直到用户需要添加iplist(addList)列表中的数据添加完才跳出循环
        while len(addList) != 0:
            for lst in config_ipPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = addList[0]
                    lst.append(addList[0])
                    addList.remove(addList[0])
                    #addIpPrefixListCommand(comlst, serverName, lst[0], addIpStr)
                    if len(addList) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(config_ipPrefixList) == True:
                tl = []
                if len(servce_ipostfix_num[serverName]) != 0:
                    postFixNum = max(servce_ipostfix_num[serverName]) + 1
                    servce_ipostfix_num[serverName].append(postFixNum)
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '_HeaderEnrich"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_' + str(postFixNum) + '_HeaderEnrich"'
                tl.append(ipliststr)
                #entryId = getTheCompatibleEntryIdByDict()
                #createIpPrefixListEntry(comlst, serverName, url, ipliststr, entryId)
                config_ipPrefixList.append(tl)


def createIpPrefixListEntry(clst, service_name, url, ip_list_str, enId):
    clst.append("exit all\n")
    clst.append("configure application-assurance group 1:1 policy\n")
    clst.append("app-filter\n")
    clst.append("entry " + str(enId) + " create\n")
    expression_1 = None
    expression_2 = None
    http_port = None
    if url != None:
        expression_1, expression_2, http_port = ChargingContextAai.processUrl(url)

    if expression_1 != None:
        clst.append('expression 1 http-host eq "' + expression_1 + '"\n')
    if expression_2 != None:
        clst.append('expression 2 http-uri eq "' + expression_2 + '"\n')

    clst.append("server-address eq " + ip_list_str + '"\n')
    clst.append('application "APP_' + service_name + '_HeaderEnrich"\n')
    clst.append("no shutdown\n")
    clst.append("exit\n")

    '''
    clst.append('exit all\n')
    clst.append('configure application-assurance group 1:1\n')
    clst.append(ip_list_str + "\n")
'''
    clst.append("\n")


def getTheCompatibleEntryIdByDict():
    global serviceEntryIdDict
    global allEntryIdList
    global allEntryIdDict

    serviceCaseStr = "head"
    retId = -1

    for i in range(2000,20000):
        if i not in allEntryIdDict[serviceCaseStr]:
            retId = i
            allEntryIdDict[serviceCaseStr].append(retId)
            break


    return retId


def addIpPrefixListCommand(clst, sName, ipStr):
    if "/" not in ipStr:
        ipStr = ipStr + "/32"
    clst.append("prefix " + ipStr + ' name "' + sName + '_HeaderEnrich"' + "\n")


def ip_prefix_list_is_full(cfg_ipPrefixList):
    global ip_prefix_list_max_number
    for lst in cfg_ipPrefixList:
        if len(lst) < ip_prefix_list_max_number + 1:
            return False
    return True


# def DeleteTheIpPrefixList(comlst, server_name, url, del_list):
#     pass
#
#
# def getTheServicePostFixNum():
#     pass
#
#
# def getTheServiceUrlIpListDict_Delete(List_del):
#     print("11111", List_del)


def gen_prefix_enrich(path, confgList):
    global configList
    configList = []
    configList = confgList
    log_list = []
    global serviceUrlIpListDict
    global serviceUrlIpListUserDict
    global del_serviceUrlIpListUserDict
    global allEntryIdList
    global serviceDict
    global serviceEntryIdDict
    # 用个全局变量来存储ip_prefix_list的最大值
    global ip_prefix_list_max_number
    ip_prefix_list_max_number = 50

    serviceUrlIpListDict = {}
    serviceUrlIpListUserDict = {}
    del_serviceUrlIpListUserDict = {}
    commandList = []
    global allEntryIdDict
    allEntryIdDict = {}
    global servicePortListDict
    servicePortListDict = {}
    global portListCommandList
    portListCommandList = []
    ccl7_cfg = open(path + '\\configureL7.log', 'r')
    ccl7_cfg_list = ccl7_cfg.readlines()
    ccl7_cfg.close()
    allEntryIdList = eval(ccl7_cfg_list[0])
    serviceDict = eval(ccl7_cfg_list[1])
    serviceEntryIdDict = eval(ccl7_cfg_list[2])
    allEntryIdDict = eval(ccl7_cfg_list[3])
    servicePortListDict = eval(ccl7_cfg_list[4])
    excel_path = path + "\\ip_prefix_list_L7_headEnrich.xlsx"
    print("9999++++",excel_path)


    global servce_ipostfix_num
    servce_ipostfix_num = {}
    global servicePostFixNumDict
    servicePostFixNumDict = {}

    excel = openpyxl.load_workbook(excel_path)
    try:
        sheet = excel["ip_prefix_list_add"]
    except Exception as err:
        print(err)

    serviceList = []
    # 该函数会根据分割符来把一条条目中包含port range的分成若干条
    serviceList = getServiceListByList(sheet, 1)

    resultList = []
    resultList = arrangeTheList(serviceList)

    resultList = arrangeTheList_2(resultList)

    for resultLine in resultList:
        if len(resultLine) == 0:
            resultList.remove(resultLine)

    for resultlst in resultList:
        addServiceUrlIpListToDict(resultlst)
    # 添加业务url列表（该业务对应的url所有的需要添加的IP）
    for resultlst in resultList:
        addServiceUrlIpListToUserDict(resultlst)

    log_list.append("这是配置文件中相关业务的数据:\n")
    for sNameKey in serviceUrlIpListDict:
        log_str = sNameKey
        for urlKey in serviceUrlIpListDict[sNameKey]:
            log_str = log_str + "  " + urlKey
            for linelst in serviceUrlIpListDict[sNameKey][urlKey]:
                log_str = log_str + "    " + str(linelst) + "\n"
                log_list.append(log_str)
                log_str = ""

    log_list.append("这是用户添加的数据:\n")
    for sNameKey in serviceUrlIpListUserDict:
        log_str = sNameKey
        for urlKey in serviceUrlIpListUserDict[sNameKey]:
            log_str = log_str + "  " + urlKey
            log_str = log_str + "    " + str(len(serviceUrlIpListUserDict[sNameKey][urlKey])) + str(
                serviceUrlIpListUserDict[sNameKey][urlKey]) + "\n"
            log_list.append(log_str)
            log_str = ""

    log_list.append("这是配置文件中相关业务的数据(添加前):\n")
    for sNameKey in serviceUrlIpListDict:
        log_str = sNameKey
        for urlKey in serviceUrlIpListDict[sNameKey]:
            log_str = log_str + "  " + urlKey
            for linelst in serviceUrlIpListDict[sNameKey][urlKey]:
                log_str = log_str + "    " + str(len(linelst) - 1), str(linelst) + "\n"
                log_list.append(log_str)
                log_str = ""

    # 先对业务进行增加操作
    #commandList.append("exit all\n")
    #commandList.append("configure application-assurance group 1:1 policy\n")
    for sNameKey in serviceUrlIpListUserDict:
        log_list.append("************"+sNameKey+str(serviceUrlIpListUserDict))
        #commandList.append("对" + sNameKey + "业务进行新增操作\n")
        for urlKey in serviceUrlIpListUserDict[sNameKey]:
            addCommandTocommandList(commandList, sNameKey, urlKey, serviceUrlIpListUserDict[sNameKey][urlKey])

    log_list.append("这是配置文件中相关业务的数据(添加后):\n")
    for sNameKey in serviceUrlIpListDict:
        log_str = sNameKey
        commandList.append("对" + sNameKey + "业务进行添加" + "\n")
        for urlKey in serviceUrlIpListDict[sNameKey]:
            log_str = log_str + "  " + urlKey
            commandList.append("   对" + urlKey + "进行操作" + "\n")
            for linelst in serviceUrlIpListDict[sNameKey][urlKey]:
                log_list.append("    " +str(len(linelst)-1)+str(linelst) + "\n")
                print("    " +str(len(linelst)-1)+str(linelst))
                for line in linelst:
                    #print("6666----",line)
                    if "create" in line:
                        commandList.append('exit all\n')
                        commandList.append('configure application-assurance group 1:1\n')
                        commandList.append(line.replace(" create", "").replace("  ", "") + "\n\n")
                    elif "ip-prefix-list" in line and "create" not in line:
                        entryId = getTheCompatibleEntryIdByDict()
                        createIpPrefixListEntry(commandList, sNameKey, urlKey, line, entryId)
                        commandList.append('exit all\n')
                        commandList.append('configure application-assurance group 1:1\n')
                        commandList.append(line.replace(" ", "") +" create" "\n\n")
                    elif "name" not in line:
                        addIpPrefixListCommand(commandList, sNameKey, line)

    text_cfg = []
    text_cfg.append(str(allEntryIdList) + "\n")
    text_cfg.append(str(serviceDict) + "\n")
    text_cfg.append(str(serviceEntryIdDict) + "\n")
    text_cfg.append(str(allEntryIdDict) + "\n")
    text_cfg.append(str(servicePortListDict) + "\n")

    file = open(path + "\\configureL7.log", "w")
    file.writelines(text_cfg)
    file.close()

    fo = open(path + "\\脚本文件\\ip_prefix_list_HeaderEnrich.txt", "w")
    fo.writelines(commandList)
    fo.close()

    if log_list:
        fo = open(path + "\\ip_prefix_list_HeaderEnrich.log", "w")
        fo.writelines(log_list)
        fo.close()

        fo = open("tmp\\ip_prefix_list_HeaderEnrich.log", "w")
        fo.writelines(log_list)
        fo.close()

