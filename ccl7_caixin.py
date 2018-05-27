import openpyxl


def getServiceListByList(sheet,startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    #firstLineServiceId = sheet.cell(row=startRow, column=serviceId_col).value
    retList = []
    retList_del = []

    for rowNumber in range(startRow,sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if changeLag == "新增":
            if portNumberL4 != None:
                if portsplit_2 in str(portNumberL4):
                    portstr = portNumberL4.split(portsplit_2)[0]
                    portL4List.append(portNumberL4.split(portsplit_2)[1])
                else:
                    pass
                if portsplit_1 in str(portstr):
                    portstr = portstr.split(portsplit_1)
                    for p in portstr:
                        portL4List.append(p)
                else:
                    portL4List.append(portstr)
                for p in portL4List:
                    retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
            else:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
            if layerLag == "L7" and serviceName == None:
                pass
        elif changeLag == "删除":
            retList_del.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))

    retList = list(set(retList))

    return  retList,retList_del

def arrangeTheList(lst):

    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList



def getTheServiceEntryList(config_List,service_name):
    retList = []

    for i in range(0,len(config_List)):
        if 'application "APP_'+service_name in config_List[i] and "create" not in config_List[i]:
            k = i
            for j in range(k,0,-1):
                if "entry" in config_List[j] and "create" in config_List[j]:
                    f = j
                    tmplist = []
                    for h in range(f,k+3):
                        tmplist.append(config_List[h].replace("\n",""))
                    break
            retList.append(tmplist)
    return retList

def getTheIpPrefixListName(cxjs_List):
    retList = []
    for entryList in cxjs_List:
        for text in entryList:
            if "server-address eq ip-prefix-list " in text:
                retList.append(text.split("server-address eq ip-prefix-list ")[1])
    return list(set(retList))

def getTheIpPrefixList(cxjs_Ip_Prefix_List_Name,config_List):
    retList = []

    for ipListName in cxjs_Ip_Prefix_List_Name:
        for i in range(0,len(config_List)):
            if "ip-prefix-list "+ipListName in config_List[i] and "create" in config_List[i]:
                k = i
                for j in range(k,len(config_List)):
                    if "exit" in config_List[j]:
                        e = j
                        break
                tmpList = []
                tmpList.append(ipListName)
                for h in range(k+1,e):
                    if "name" in config_List[h]:
                        tmpList.append(config_List[h].split("prefix ")[1].split(" name")[0].replace("/32",""))
                retList.append(tmpList)

    return retList


def processTheCaixinDict(caixin_Dict,s_lst):
    service_name = s_lst[0][3]
    for tup in s_lst:
        if tup[7] != None:
            l7_str = tup[7].replace(":*","").replace("/*","")
            if "/" in l7_str and l7_str.split("/")[0] == l7_str.split("/")[0].upper():
                l7_str = l7_str.split("/")[0]
            if l7_str == l7_str.upper():
                caixin_Dict[service_name][0].append(l7_str)
            else:
                caixin_Dict[service_name][1].append(l7_str)
    if len(caixin_Dict[service_name][0]) != 0:
        tmplist = list(set(caixin_Dict[service_name][0]))
        tmplist.sort()
        caixin_Dict[service_name][0] = tmplist

def ip_prefix_list_is_full(cfg_ipPrefixList):
    global ip_prefix_list_max_number
    for lst in cfg_ipPrefixList:
        if len(lst) < ip_prefix_list_max_number + 1:
            return False
    return True

def addIpPrefixListCommand(clst,sName,ipStr):
    if "/" not in ipStr:
        ipStr = ipStr +"/32"
    clst.append("prefix "+ipStr+' name "'+sName+'"'+"\n")


def createTheCaixinIpPrefixList(serviceName,caixinListAdd,command_List,cxjs_IpPrefixList):
    # 若配置列表为空则表示配置文件中无该业务的ip_prefix_list
    ipstrDict = {}
    #存放彩信接受ipprefixlist列表名的后缀
    postFixNum = -1
    if len(cxjs_IpPrefixList) == 0:
        tl = []
        postFixNum = 1
        if postFixNum < 10:
            ipliststr = 'ip-prefix-list "app_' + serviceName + '_0' + str(postFixNum) + '"'
        else:
            ipliststr = 'ip-prefix-list "app_' + serviceName + '_' + str(postFixNum) + '"'
        tl.append(ipliststr)
        cxjs_IpPrefixList.append(tl)

        while len(caixinListAdd) != 0:
            for lst in cxjs_IpPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    #if lst[0] not in ipstrDict:
                    #    command_List.append('exit all\n')
                    #    command_List.append('configure application-assurance group 1:1\n')
                    #    command_List.append(lst[0] + "\n")
                    #    ipstrDict[lst[0]] = True
                    addIpStr = caixinListAdd[0]
                    lst.append(caixinListAdd[0])
                    #addIpPrefixListCommand(command_List, serviceName,lst[0], addIpStr)
                    caixinListAdd.remove(caixinListAdd[0])
                    if len(caixinListAdd) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(cxjs_IpPrefixList) == True:
                tl = []
                postFixNum+=1
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serviceName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serviceName + '_' + str(postFixNum) + '"'
                tl.append(ipliststr)
                cxjs_IpPrefixList.append(tl)

    else:
        # 循环直到用户需要添加iplist(addList)列表中的数据添加完才跳出循环
        while len(caixinListAdd) != 0:
            for lst in cxjs_IpPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    #if lst[0] not in ipstrDict:
                    #    command_List.append('exit all\n')
                    #    command_List.append('configure application-assurance group 1:1\n')
                    #    command_List.append(lst[0] + "\n")
                    #    ipstrDict[lst[0]] = True
                    addIpStr = caixinListAdd[0]
                    lst.append(caixinListAdd[0])
                    #addIpPrefixListCommand(command_List, serviceName,lst[0], addIpStr)
                    caixinListAdd.remove(caixinListAdd[0])
                    if len(caixinListAdd) == 0:
                        break

            # 若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(cxjs_IpPrefixList) == True:
                tl = []
                postFixNum = len(cxjs_IpPrefixList)+1
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serviceName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serviceName + '_' + str(postFixNum) + '"'
                tl.append(ipliststr)
                cxjs_IpPrefixList.append(tl)

def getTheCompatibleEntryIdByDict(caixin_list):
    retId = -1
    caixin_list.sort()
    for i in range(20001,20500):
        if i not in caixin_list:
            retId = i
            caixin_list.append(retId)
            break

    return retId

def getTheIpPrefixListStr(ip_address,IpPrefixList_config):
    for lst in IpPrefixList_config:
        for text in lst:
            if ip_address == text:
                return lst[0]
    return "没有找到相应的ip_prefix_list"

def createTheCaixinEntry(comLst,tup,cxjs_IpPrefixList_config,caixin_entry_id_list):
    serviceName = tup[3]
    ipStr = tup[7]
    if ipStr != None:
        ipAddress = ipStr.replace("/*", "").replace(":*", "")
        if "/" in ipAddress:
            ipAddress = ipAddress.split("/")[0]
        ipPrefixlistStr = getTheIpPrefixListStr(ipAddress, cxjs_IpPrefixList_config)
        comLst.append("exit all\n")
        comLst.append("configure application-assurance group 1:1 policy\n")
        comLst.append("app-filter\n")
        comLst.append("entry " + str(getTheCompatibleEntryIdByDict(caixin_entry_id_list)) + " create\n")
        comLst.append('expression 1 http-host eq "^' + ipStr + '$"' + "\n")
        #comLst.append("server-address eq ip-prefix-list " + ipPrefixlistStr + "\n")
        if ipPrefixlistStr == "没有找到相应的ip_prefix_list" and serviceName == "cxfs_01":
            comLst.append('server-address eq dns-ip-cache "TrustedCache"' + "\n")
        else:
            comLst.append("server-address eq " + ipPrefixlistStr + "\n")
        comLst.append('application "APP_' + serviceName + '"\n')
        comLst.append("no shutdown\n")
    else:
        comLst.append("exit all\n")
        comLst.append("configure application-assurance group 1:1 policy\n")
        comLst.append("app-filter\n")
        comLst.append("entry " + str(getTheCompatibleEntryIdByDict(caixin_entry_id_list)) + " create\n")
        comLst.append("ip-protocol-num eq " + tup[5] + "\n")
        if "/" not in tup[4]:
            comLst.append("server-address eq " + tup[4] + "/32\n")
        else:
            comLst.append("server-address eq " + tup[4] + "\n")
        if tup[6] != None:
            comLst.append("server-port eq " + tup[6] + "\n")
        comLst.append('application "APP_' + serviceName + '"\n')
        comLst.append("no shutdown\n")


def getCaixinEntryIdList(config_list):
    retList = []

    for text in config_list:
        if "entry" in text and "create" in text:
            id = int(text.split("entry ")[1].split(" create")[0])
            if id >=20001 and id <=20500:
                retList.append(id)
    return retList



def gen_caixin(configList,excel_path):
    #用个全局变量来存储ip_prefix_list的最大值
    global ip_prefix_list_max_number
    ip_prefix_list_max_number = 50
    commandList = []
    excel = openpyxl.load_workbook(excel_path+'\\内容计费整理L7_caixin.xlsx')

    sheet = excel["L7"]

    cxfsList = []
    cxfsList = getTheServiceEntryList(configList, "cxfs")
    cxjsList = []
    cxjsList = getTheServiceEntryList(configList, "cxjs")
    cxfsIpPrefixList = []
    cxjsIpPrefixList = []
    cxjsIpPrefixListName = []
    cxjsIpPrefixListName = getTheIpPrefixListName(cxjsList)
    if len(cxjsIpPrefixListName)!= 0:
        cxjsIpPrefixListName.sort()
        cxjsIpPrefixList = getTheIpPrefixList(cxjsIpPrefixListName,configList)

    serviceList = []
    # 该函数会根据分割符来把一条条目中包含port range的分成若干条
    serviceList,serviceList_del = getServiceListByList(sheet, 3)

    resultList = []
    resultList = arrangeTheList(serviceList)



    caixinDict = {}
    for lst in resultList:
        #print("---",len(lst),lst)
        caixinDict[lst[0][3]] = [[],[]]
        processTheCaixinDict(caixinDict,lst)

    # print("这是添加前的数据：")
    # print("彩信接受局数据表:",caixinDict["cxjs_01"])
    # print("配置数据:",cxjsIpPrefixList)
    front_cxjsIpPrefixList = []
    for lst in cxjsIpPrefixList:
        tmplist = []
        for t in lst:
            tmplist.append(t)
        front_cxjsIpPrefixList.append(tmplist)

    for key in caixinDict:
        if key == "cxfs_01":
            pass
        if key == "cxjs_01":
            createTheCaixinIpPrefixList(key,caixinDict[key][0],commandList,cxjsIpPrefixList)
    back_cxjsIpPrefixList = []
    for lst in cxjsIpPrefixList:
        tmplist = []
        for t in lst:
            tmplist.append(t)
        back_cxjsIpPrefixList.append(tmplist)
    com_ip_list = []

    for f_lst in front_cxjsIpPrefixList:
        iplistStr = f_lst[0]
        for b_lst in back_cxjsIpPrefixList:
            if iplistStr in b_lst:
                for f_lint in f_lst:
                    b_lst.remove(f_lint)
                b_lst.insert(0,iplistStr)


    # print("这是添加后的数据：")
    # print("彩信接受局数据表:", caixinDict["cxjs_01"])
    # print("配置数据:", cxjsIpPrefixList)

    for lst in back_cxjsIpPrefixList:
        for text in lst:
            if "app_" in text:
                commandList.append('\nexit all\n')
                commandList.append('configure application-assurance group 1:1\n')
                commandList.append(text + "\n")
            else:
                addIpPrefixListCommand(commandList, "cxjs_01", text)

    caixinEntryIdList = []

    caixinEntryIdList = getCaixinEntryIdList(configList)
    commandList.append("\n\n")
    for lst in resultList:
        if lst[0][3]=="cxjs_01":
            for tup in lst:
                createTheCaixinEntry(commandList,tup,cxjsIpPrefixList,caixinEntryIdList)
        if lst[0][3]=="cxfs_01":
            for tup in lst:
                createTheCaixinEntry(commandList,tup,cxfsIpPrefixList,caixinEntryIdList)

    fo = open(excel_path+"\\caixin_ip_prefix_list.txt", "w")
    fo.writelines(commandList)
    fo.close()