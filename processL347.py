# coding:utf-8
import os

import openpyxl
import ChargingContextAai
import ccL7
import ccl34
import ccl7_HeaderEnrich
import ccl7_ip_prefix_list_HeaderEnrich
import ccl7_iplist
import ccl7_iplist_del
import full_gen
import time
import zipfile,ccl7_caixin,specialServiceIpPrefixList


def getServiceListByList(sheet, startRow):
    global serviceCaseList
    global head_enrich_list
    global caixin_list
    global L3_in_L7_NameList
    global L3_in_L7_List
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    HeaderEnrich_col = 19
    serviceCase_col = 9
    layerLag = ""
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        if ipAddressL3 !=None:
            if " " in ipAddressL3:
                ipAddressL3 = ipAddressL3.replace(" ","")
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        #if urlL7 !=None and "https://" in urlL7 and portNumberL4 == None:
         #       portNumberL4 = 443
        if portNumberL4 != None:
            portNumberL4 = str(portNumberL4)
        if serviceId !=None and serviceName !=None:
            if urlL7 != None and "." not in urlL7:
                urlL7 = None
            HeaderEnrich = sheet.cell(row=rowNumber, column=HeaderEnrich_col).value
            if HeaderEnrich != None:
                if "头增强" in HeaderEnrich:
                    head_enrich_list.append(
                        (changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
            #print("99999++",changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7)
            if "cxjs" in serviceName or "cxfs" in serviceName:
                caixin_list.append((changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
            elif serviceName in L3_in_L7_NameList and ipAddressL3 !=None and protocolNumber==None and portNumberL4==None and urlL7 == None:
                #print("6666666++++++",changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7)
                L3_in_L7_List.append((changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
            else:
                retList.append((changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    return retList


def arrangeTheList(lst, configureList):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(str(serviceId))
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    newRetList = []
    tlst = []
    for retline in retList:
        layerLag = selectL347lag(retline, configureList)
        for tup in retline:
            changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            tlst.append((layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url))
        newRetList.append(tlst)
        tlst = []
    return newRetList


def selectL347lag(tupList, cfgLst):
    global log_list
    for i in range(0, len(cfgLst)):
        try:
            if 'policy-rule-unit "PRU_' + tupList[0][2] in cfgLst[i] and "qci * arp * precedence" not in cfgLst[i]:
                k = i
                for j in range(k, len(cfgLst)):
                    if "aa-charging-group" in cfgLst[j]:
                        log_list.append("该业务" + tupList[0][2] + "是L7")
                        return "L7"
                    if "protocol" in cfgLst[j]:
                        log_list.append("该业务" + tupList[0][2] + "是L4")
                        return "L4"
                    if "remote-ip" in cfgLst[j] and "protocol" not in cfgLst[j - 1]:
                        log_list.append("该业务" + tupList[0][2] + "是L3")
                        return "L3"
                    if "exit" in cfgLst[j]:
                        break
        except:
            pass

    for tup in tupList:
        if tup[6] != None:
            log_list.append("该业务" + tupList[0][2] + "是L7为新业务\n")
            return "L7"
    for tup in tupList:
        if tup[4] != None or tup[5] != None:
            log_list.append("该业务" + tupList[0][2] + "是L4为新业务\n")
            return "L4"
    return "L3"


def writeRowInExcel(sheet, px, py, writetup):
    sheet.cell(row=py, column=1, value=writetup[0])
    sheet.cell(row=py, column=3, value=writetup[1])
    sheet.cell(row=py, column=8, value=writetup[2])
    sheet.cell(row=py, column=10, value=writetup[3])
    sheet.cell(row=py, column=13, value=writetup[4])
    sheet.cell(row=py, column=14, value=writetup[5])
    sheet.cell(row=py, column=15, value=writetup[6])
    sheet.cell(row=py, column=16, value=writetup[7])


def getUrlTimes(tups7, urlstr,portstr):
    t = 0
    for line in tups7:
        if line[7] == urlstr and line[6]==portstr:
            t += 1
    return t


def isIpList(serviceName, cfglist, httphost,serviceport):
    http_host,uri,port = ChargingContextAai.processUrl(httphost)
    #print("host+++",http_host,uri,serviceport)
    for i in range(0, len(cfglist)):
        if http_host in cfglist[i]:  # 先判断http_host 再判断下面的server-address eq ....来确认是否是iplist
            k = i
            for j in range(k,len(cfglist)):
                if "no shutdown" in cfglist[j]:
                    e = j
                    if ChargingContextAai.EntryIsTrue(serviceName,serviceport,http_host,uri,port,cfglist,k-1,e) ==True:
                        for t in range(k,e):
                            if 'server-address eq ip-prefix-list' in cfglist[t]:
                                return True
                    break

    return False


def getIPlistServiceTupList(tupLst7, configList):
    global log_list
    L7list = []
    urlportList = []
    # url_times = []
    for tup in tupLst7:
        if tup[7] != None:
            L7list.append(tup)
            urlportList.append((tup[7],tup[6]))
    urlportList = list(set(urlportList))
    serviceDict = {}
    iplist = []
    #print("L7的条目数为:",len(L7list),L7list)
    for tup in L7list:
        if isIpList(tup[3], configList, tup[7],tup[6]) == True:
            iplist.append(tup)
    #print("++++",urlportList)
    for url,port in urlportList:
        urlportTime = 0
        if url != None:
            urlportTime = getUrlTimes(tupLst7, url,port)
        #print(str(url)+str(port) + "出现次数:" + str(urlportTime))
        log_list.append(str(url)+str(port) + "出现次数:" + str(urlportTime) + "\n")
        if urlportTime > 5:
            #print("该" + str(url)+str(port) + "出现次数超过5次应放入ip-prefix-list:" + str(urlportTime))
            log_list.append("该" + str(url)+str(port) + "出现次数超过5次应放入ip-prefix-list:" + str(urlportTime) + "\n")
            for tup in tupLst7:
                if tup[7] == url and tup[6] == port:
                    iplist.append(tup)
    return list(set(iplist))


def writeExcel(lst, postfix, configList, path):
    global log_list
    tupListL34 = []
    tupListL7 = []
    for llst in lst:
        if llst[0][0] == "L3":
            for tup in llst:
                tupListL34.append(tup)

    for llst in lst:
        if llst[0][0] == "L4":
            for tup in llst:
                tupListL34.append(tup)

    for llst in lst:
        if llst[0][0] == "L7":
            for tup in llst:
                tupListL7.append(tup)
    fPath = None
    if len(tupListL34) != 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "L34"
        x = 1
        y = 1
        for tup in tupListL34:
            log_list.append(str(tup) + "该条目被存入‘内容计费整理L34’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\内容计费整理L34" + ".xlsx"
        else:
            fPath = path + "\\内容计费整理L34" + postfix + ".xlsx"
        if fPath != None:
            wb.save(fPath)
            wb.close()

    fPath = None
    ip_list_tupList = []
    if len(tupListL7) != 0:
        ip_list_tupList = getIPlistServiceTupList(tupListL7, configList)
        for line in ip_list_tupList:
            tupListL7.remove(line)

    if len(tupListL7) != 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "L7"
        x = 1
        y = 1
        for tup in tupListL7:
            log_list.append(str(tup) + "该条目被存入‘内容计费整理L7’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\内容计费整理L7" + ".xlsx"
        else:
            fPath = path + "\\内容计费整理L7" + postfix + ".xlsx"
        if fPath != None:
            wb.save(fPath)
            wb.close()
    fPath = None
    del_ip_list_tupList = []
    add_ip_list_tupList = []
    for tup in ip_list_tupList:
        if tup[1] == "删除":
            del_ip_list_tupList.append(tup)
        if tup[1] == "新增":
            add_ip_list_tupList.append(tup)

    if len(add_ip_list_tupList) != 0 or len(del_ip_list_tupList) != 0:
        if len(add_ip_list_tupList) != 0:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "ip_prefix_list_add"
        if len(del_ip_list_tupList) != 0:
            sheet_del = wb.create_sheet("ip_prefix_list_del")
    if len(add_ip_list_tupList) != 0:
        x = 1
        y = 1
        for tup in add_ip_list_tupList:
            log_list.append(str(tup) + "该条目被存入‘ip_prefix_list_L7EXCEL的ip_prefix_list_add’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\ip_prefix_list_L7" + ".xlsx"
        else:
            fPath = path + "\\ip_prefix_list_L7" + postfix + ".xlsx"

    if len(del_ip_list_tupList) != 0:
        x = 1
        y = 1
        for tup in del_ip_list_tupList:
            log_list.append(str(tup) + "该条目被存入‘ip_prefix_list_L7EXCEL的ip_prefix_list_del’表格中" + "\n")
            writeRowInExcel(sheet_del, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\ip_prefix_list_L7" + ".xlsx"
        else:
            fPath = path + "\\ip_prefix_list_L7" + postfix + ".xlsx"
    try:
        if fPath != None:
            wb.save(fPath)
            wb.close()
    except:
        pass
    return tupListL34, tupListL7, del_ip_list_tupList, add_ip_list_tupList


def getAllEntryIdDict(all_entry_id_dict, all_entry_id_list):
    list_head = []
    list_no_head = []
    list_white = []
    list_caixin = []
    for id in all_entry_id_list:
        if id >= 2001 and id < 20000:
            list_head.append(id)
        if id >= 20000 and id < 20501:
            list_caixin.append(id)
        if id >= 20501 and id < 60000:
            list_no_head.append(id)
        if id >= 50000 and id < 65000:
            list_white.append(id)

    all_entry_id_dict["head"] = list_head
    all_entry_id_dict["caixin"] = list_caixin
    all_entry_id_dict["no_head"] = list_no_head
    all_entry_id_dict["white"] = list_white


def getAllEntryIdList(all_entry_list, cfglst):
    for i in range(0, len(cfglst)):
        if 'entry' in cfglst[i] and "create" in cfglst[i]:
            all_entry_list.append(int(cfglst[i].split("entry ")[1].split(" create")[0]))




def gen_origin_api(*args):
    serviceDi = []
    path = ''
    flag=True
    global serviceDict
    serviceDict = {}
    global serviceEntryIdDict
    serviceEntryIdDict = {}
    global allEntryIdList
    global allEntryIdDict
    global log_list
    log_list = []
    global caixin_list
    caixin_list = []
    global L3_in_L7_NameList
    L3_in_L7_NameList = ["txsp_00"]
    global L3_in_L7_List
    L3_in_L7_List = []
    path = os.path.abspath('.')
    try:
        configFile = open(args[1], 'r')
        configList = configFile.readlines()
        configFile.close()
    except Exception:
        configFile = open(args[1], 'r',encoding='utf-8')
        configList = configFile.readlines()
        configFile.close()
    global head_enrich_list
    head_enrich_list = []
    global serviceCaseList
    serviceCaseList = []
    global servicePortListDict
    servicePortListDict = {}
    allPortList = []
    allEntryIdList = []
    getAllEntryIdList(allEntryIdList, configList)
    # 获取所有entryId(免统定收【头增强】白)存入字典
    allEntryIdDict = {}
    getAllEntryIdDict(allEntryIdDict, allEntryIdList)
    text_cfg = []
    text_cfg.append(str(allEntryIdList) + "\n")
    text_cfg.append(str(serviceDict) + "\n")
    text_cfg.append(str(serviceEntryIdDict) + "\n")
    text_cfg.append(str(allEntryIdDict) + "\n")
    text_cfg.append(str(servicePortListDict) + "\n")

    l_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    for ne_name in configList:
        if ne_name.find('name "') != -1 and ne_name.find('BNK"') != -1:
            start = ne_name.find('"')
            end = ne_name.find('"', start + 1)
            ne_name = ne_name[start+1:end]
            path = path + '\\' + 'Generated\\' + ne_name + '\\' + l_time + '\\' + args[0][0:-5].replace('tmp\\', '')
            mkdir(path + "\\脚本文件")
            flag=False
            break
    if flag:
        path = path + '\\' + 'Generated\\' + 'unknownNE' + '\\' + l_time + '\\' + args[0][0:-5].replace('tmp\\', '')
        #mkdir(path)
        mkdir(path+"\\脚本文件")
    full_gen.gen_begining(configList, path)
    file = open(path + "\\configureL7.log", "w")
    file.writelines(text_cfg)
    file.close()
    excel = openpyxl.load_workbook(args[0])
    sheet = excel["本次变更"]
    serviceList = []
    serviceList = getServiceListByList(sheet, 3)


    resultList = arrangeTheList(serviceList, configList)

    resultList_head = arrangeTheList(head_enrich_list, configList)
    statistics_list = writeExcel(resultList, None, configList, path)

    if os.path.exists(path + "\\内容计费整理L34.xlsx"):
        ccl34.gen_l34(configList, path)
    # else:
    #     fo_log = open("L34.log", "w")
    #     fo_log.writelines(['本次无34层数据变更'])
    #     fo_log.close()
    if L3_in_L7_List:
        L3_in_L7_List_result = arrangeTheList(L3_in_L7_List, configList)
        writeExcel(L3_in_L7_List_result, "_specialService", configList, path)
        specialServiceIpPrefixList.gen_spec(configList,path)
    if os.path.exists(path + "\\内容计费整理L7.xlsx"):
        serviceDi = ccL7.gen_l7(configList, path)
    # else:
    #     fo_log = open("L7.log", "w")
    #     fo_log.writelines(['本次无7层数据变更'])
    #     fo_log.close()
    if os.path.exists(path + "\\ip_prefix_list_L7.xlsx"):
        ccl7_iplist.gen_iplist(configList, path)
        ccl7_iplist_del.gen_iplist_del(configList, path)
    if resultList_head:
        writeExcel(resultList_head, "_headEnrich", configList, path)
        if os.path.exists(path + "\\内容计费整理L7_headEnrich.xlsx"):
            ccl7_HeaderEnrich.gen_hearderenrich(path, configList)
    if os.path.exists(path + "\\ip_prefix_list_L7_headEnrich.xlsx"):
        ccl7_ip_prefix_list_HeaderEnrich.gen_prefix_enrich(path, configList)

    if caixin_list:
        caixin_list_result = arrangeTheList(caixin_list, configList)
        writeExcel(caixin_list_result, "_caixin", configList, path)
        ccl7_caixin.gen_caixin(configList,path)



    fo = open(path + "\\processL347.log", "w")
    fo.writelines(log_list)
    fo.close()

    zipfile = path + '.zip'
    loc_fo = open("tmp\\processL347.log", "w")
    loc_fo.writelines(log_list)
    loc_fo.close()
    zip_ya(path, zipfile)


def mkdir(path):
    path = path.strip()
    isExists = os.path.exists(path)
    if not isExists:
        os.makedirs(path)
        return True
    else:
        return False


def zip_ya(startdir, file_news):
    z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)  # 参数一：文件夹名
    for dirpath, dirnames, filenames in os.walk(startdir):
        fpath = dirpath.replace(startdir, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        fpath = fpath and fpath + os.sep or ''  # 这句话理解我也点郁闷，实现当前文件夹以及包含的所有文件的压缩
        for filename in filenames:
            z.write(os.path.join(dirpath, filename), fpath + filename)
            # print('压缩成功')
    z.close()
