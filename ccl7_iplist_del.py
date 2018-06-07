import openpyxl


def getServiceListByList(sheet, startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if portNumberL4 != None:
            if portsplit_2 in str(portNumberL4):
                portstr = portNumberL4.split(portsplit_2)[0]
                portL4List.append(portNumberL4.split(portsplit_2)[1])
            else:
                pass
            if portsplit_1 in str(portstr):
                portstr = portstr.split(portsplit_1)
                for p in portstr:
                    portL4List.append(p)
            else:
                portL4List.append(portstr)
            for p in portL4List:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
        else:
            retList.append(
                (layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    retList = list(set(retList))

    return retList


def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList


def getTheUrlList(lst):
    retList = []
    for tup in lst:
        retList.append(tup[7])
    retList = list(set(retList))
    return retList


def getTheTupList(url, sLst):
    retList = []
    for tup in sLst:
        if tup[7] == url:
            retList.append(tup)
    return retList


def getTheServiceUrlIpListDict_Delete(List_del):
    retDict = {}

    for serviceList in List_del:
        serviceName = serviceList[0][3]
        urlList = getTheUrlList(serviceList)
        urlDict = {}
        for url in urlList:
            tupList = getTheTupList(url, serviceList)
            urlDict[url] = tupList
            retDict[serviceName] = urlDict

    return retDict


def getTheiplistStrList(listName, clst):
    retList = []
    for i in range(0, len(clst)):
        if 'ip-prefix-list ' + listName in clst[i] and "create" in clst[i]:
            k = i
            for j in range(k, len(clst)):
                if "exit" in clst[j]:
                    retList.append(clst[j])
                    break
                else:
                    retList.append(clst[j])
    return retList


def getTheIpPrefixListStrList(sName, urlstr, cfglst):
    retList = []
    iplist_name_list = []

    http_host = urlstr.replace("http://", "").replace("/*", "").replace(":*", "")
    for i in range(0, len(cfglst)):
        if 'expression 1 http-host eq' in cfglst[i] and http_host in cfglst[i]:
            k = i
            for j in range(k, len(cfglst)):
                if 'application "APP_' + sName + '"' in cfglst[j]:
                    if 'server-address eq ip-prefix-list' in cfglst[j - 1]:
                        prefixlistname = cfglst[j - 1].split("server-address eq ip-prefix-list ")[1]
                        iplist_name_list.append(prefixlistname.replace("\n", ""))
                    break

    for iplistName in iplist_name_list:
        iplistStrList = getTheiplistStrList(iplistName, cfglst)
        retList.append(iplistStrList)
    return retList


def getTheServiceUrlIpListStrDict_Delete(service_dict, cfglst):
    retDict = {}
    for service_name in service_dict:
        urlDict = {}
        for url in service_dict[service_name]:
            iplistStr = []
            iplistStr = getTheIpPrefixListStrList(service_name, url, cfglst)
            urlDict[url] = iplistStr
            retDict[service_name] = urlDict

    return retDict


def getTheDeleteIpPrefixListName(tup, strList):
    retStr = ""
    ipStr = tup[4].replace(" ", "")
    if "/" not in ipStr:
        ipStr = ipStr + "/32"
    for i in range(0, len(strList)):
        # k = i
        for text in strList[i]:
            if ipStr in text:
                retStr = "ip-prefix-list " + strList[i][0].split("ip-prefix-list ")[1].split(" create")[0]
                return retStr

    return str(tup) + "未找到该条目"


def getTheSplitList(tempList):
    nameList = []
    retList = []
    for c in tempList:
        nameList.append(c[0])
    nameList = list(set(nameList))
    for name in nameList:
        tList = []
        tList.append(name)
        for c in tempList:
            if c[0] == name:
                tList.append(c[1])
        retList.append(tList)
    return retList


def DeleteTheUrlIpList(comlst, service_name, url, delete_tup_list, ipListStrList):
    global log_list
    comlst.append("对" + service_name + "的" + url + "进行操作\n")
    comlst.append("exit all\n")
    comlst.append("configure application-assurance group 1:1\n")
    tempList = []
    for tup in delete_tup_list:
        deleteName = getTheDeleteIpPrefixListName(tup, ipListStrList)
        if "未找到该条目" not in deleteName:
            ipStr = tup[4].replace(" ", "")
            if "/" not in ipStr:
                ipStr = ipStr + "/32"
                tempList.append((deleteName, ipStr))
            else:
                tempList.append((deleteName, ipStr))
    SplitList = []
    SplitList = getTheSplitList(tempList)
    log_list.append("对" + service_name + "的" + url + "进行操作\n")
    for lst in SplitList:
        for line in lst:
            if "ip-prefix-list" in line:
                comlst.append(line + "\n")
                log_list.append("进入该ip-prefix-list："+line + "\n")
            else:
                comlst.append("no prefix " + line + "\n")
                log_list.append("删除prefix：" + line + "\n")
        comlst.append("\n")


def DeleteTheServiceName_iplist(comlist, sName, delete_url_dict, ipListStrDict):
    for url_key in delete_url_dict:
        DeleteTheUrlIpList(comlist, sName, url_key, delete_url_dict[url_key], ipListStrDict[sName][url_key])


def gen_iplist_del(configList, path):
    global log_list
    log_list = []
    commandList = []
    excel_path = path + "\\ip_prefix_list_L7.xlsx"
    excel = openpyxl.load_workbook(excel_path)
    codeFlag = True
    try:
        sheet_del = excel["ip_prefix_list_del"]
    except Exception as err:
        print(err)
        codeFlag = False
    if codeFlag == True:
        # 对ip_prefix_list进行删除操作
        resultList_del = getServiceListByList(sheet_del, 1)
        print("iplist del is",resultList_del)
        resultList_del = arrangeTheList(resultList_del)

        # 获取{业务名：{url:[ip]}}该结构的字典，删除操作
        serviceUrlIpListDict_Delete = {}
        serviceUrlIpListDict_Delete = getTheServiceUrlIpListDict_Delete(resultList_del)

        # 获取业务名对应的url对应的iplist列表
        # {业务名：{url:[[iplist的字符段]]}}
        serviceUrlIpListStrDict_Delete = {}
        serviceUrlIpListStrDict_Delete = getTheServiceUrlIpListStrDict_Delete(serviceUrlIpListDict_Delete, configList)
        for key in serviceUrlIpListStrDict_Delete:
            log_list.append(str(key)+str(serviceUrlIpListDict_Delete[key])+"\n")
            for urlKey in serviceUrlIpListStrDict_Delete[key]:
                log_list.append(str(key) +str(urlKey)+ str(serviceUrlIpListStrDict_Delete[key][urlKey]) + "\n")

        # 进行删除操作
        for del_serviceName_key in serviceUrlIpListDict_Delete:
            DeleteTheServiceName_iplist(commandList, del_serviceName_key, serviceUrlIpListDict_Delete[del_serviceName_key],
                                        serviceUrlIpListStrDict_Delete)

        if commandList:
            fo = open(path + "\\脚本文件\\ip_prefix_list_del.txt", "w")
            fo.writelines(commandList)
            fo.close()

        if log_list:
            fo_log = open(path + "\\ip_prefix_list_del_log", "w")
            fo_log.writelines(log_list)
            fo_log.close()

            fo_log = open("tmp\\ip_prefix_list_del.log", "w")
            fo_log.writelines(log_list)
            fo_log.close()

