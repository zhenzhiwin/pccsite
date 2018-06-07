import gc
import os
import time

import openpyxl

import processL347


class pRCD():
    def __init__(self, type, msisdn, imsi, apn, rat, st, et, rg):
        m = ''
        im = ''
        self.type = type
        for i in range(0, len(msisdn), 2):
            m = m + msisdn[i + 1] + msisdn[i]
        self.msisdn = m[0:-1]
        for i in range(0, len(imsi), 2):
            im = im + imsi[i + 1] + imsi[i]
        self.imsi = imsi[0:-1]
        self.apn = apn
        self.rat = rat
        self.st = st
        self.et = et
        self.rg = rg


class SubPRCD():
    def __init__(self, rg_st, rg_et, upvol, downvol, sid):
        self.rg_st = rg_st
        self.rg_ed = rg_et
        self.upvol = upvol
        self.downvol = downvol
        self.sid = sid


def get_tup(loglist):
    front = -1
    end = -1
    lst = []
    for a in range(0, len(loglist)):
        if ">pGWRecord(79)" in loglist[a]:
            front = a
            for j in range(front, len(loglist)):
                if "consolidationResult" in loglist[j]:
                    end = j
                    lst.append((front, end))
                    break
    return lst


def getWholeChargingListByfile(loglist):
    front = -1
    end = -1
    lst = []
    for a in range(0, len(loglist)):
        # print(loglist[a])
        try:
            if ("]#" in loglist[a] or "]$" in loglist[a]) and ">pGWRecord(79)" in loglist[a + 1]:
                front = a
        except:
            pass
        try:
            if "consolidationResult" in loglist[a] and ("]#" in loglist[a + 2] or "]$" in loglist[a + 2]):
                end = a
                lst.append((front, end))
                front = -1
                end = -1
        except:
            pass

        # loglist[a]
    # print(lst)
    return lst


def getRecordList(lst, front, end):
    allrecordlist = []
    ltemp = []
    # if ">" in lst[199]:
    #   print("recode")
    # if "101" in lst[69] or "consolidationResult" in lst[i]:
    #   print("recode")
    for i in range(front, end):
        if ">" in lst[i]:
            # print(lst[i])
            ltemp.append(lst[i])
            # consolidationResult
        # if "consolidationResult" in lst[i]:
        # print(lst[i])
        # if "101" in lst[i] or "consolidationResult" in lst[i]:
        if "consolidationResult" in lst[i]:
            # print("+++++"+str(i))
            # ltemp.append(lst[i])
            allrecordlist.append(ltemp)
            ltemp = []
            # print("有空格"+lst[i])
            # prin(i)
        # print(ltemp)
        # print(allrecordlist)
        # time.sleep(1)
        # print("======"+str(i))
    # print(allrecordlist)
    return allrecordlist


def getBillFileName(billstr):
    # billstr.split(" ")[3].replace("/tmp","").replace("/","").replace(".res","").replace(".txt","")
    # print(billstr)
    for line in billstr.split(" "):
        if "cg" in line and ".dat" in line:
            # print(line)
            return line.replace("/tmp", "").replace("/", "").replace(".res", "").replace(".txt", "")


def write07Excel(fname, dictlist):
    path = os.path.abspath('.')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "text"
    titlelist = ["序号", "话单类型", "测试号码", "IMSI", "APN", "RAT Type", "startTime", "stopTime", "RG开始时间", "RG结束时间", "上行流量",
                 "下行流量", "SERVICE_ID", "话单文件", "CG IP地址"]
    x = 1
    for v in titlelist:
        sheet.cell(row=1, column=x, value=v)
        x += 1
    # y=1
    for key in dictlist:
        # print(key)
        for keyValueList in dictlist[key]:
            writeRowInExcel(sheet, key, keyValueList)
            # y+=1
            # print(keyValueList)
    l_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    path = path + '\\' + 'Generated\\' + '\\计费验证\\' + l_time
    processL347.mkdir(path)
    path = path + '\\' + l_time + '.xlsx'
    wb.save(path)
    wb.close()

    # print("写入数据成功！")


def getnumber(str):
    # print(str)
    retstr = ""
    for n in range(0, len(str), 2):
        if n % 2 == 0:
            tstr = str[n + 1] + str[n]
            retstr += tstr
    # print(retstr.replace("f",""))
    return retstr.replace("f", "")


def getRGlist(lst):
    rtlist = []
    tmplist = []
    for text in lst:
        if "serviceIdentifier" in text:
            tmplist.append(text.split("value:")[1])
            rtlist.append(tmplist)
            tmplist = []
        else:
            # 获取rg开始时间
            if "timeOfFirstUsage" in text:
                tmplist.append(text.split("value:")[1])
            if "timeOfLastUsage" in text:
                tmplist.append(text.split("value:")[1])
            if "datavolumeFBCUplink" in text:
                tmplist.append(text.split("value:")[1])
            if "datavolumeFBCDownlink" in text:
                tmplist.append(text.split("value:")[1])
    # print(rtlist)
    return rtlist


def handleOneRecord(wlist):
    # 该函数把需要填的验证内容存入列表
    rList = []
    rglist = []
    retList = []
    # 获取话单类型
    for text in wlist:
        if "recordType" in text:
            rList.append(text.split(":")[1])
            break
    # 获取msisdn
    for text in wlist:
        if "servedMSISDN" in text:
            numstr = text.split(":")[1].replace(" 0x", "")
            msisdn = getnumber(numstr)
            rList.append(msisdn)
            break
    # 获取imsi
    for text in wlist:
        if "servedIMSI" in text:
            numstr = text.split(":")[1].replace(" 0x", "")
            imsi = getnumber(numstr)
            rList.append(imsi)
            break
    # 获取APN
    for text in wlist:
        if "accessPointNameNI" in text:
            rList.append(text.split(":")[1])
            break

    # 获取基站类型
    for text in wlist:
        if "rATType" in text:
            rList.append(text.split(":")[1])
            break
    # 获取记录开始时间
    for text in wlist:
        if "recordOpeningTime" in text:
            rList.append(text.split("value:")[1])
            break
    # 获取记录结束时间(stoptime)
    for text in wlist:
        if "timeOfReport" in text:
            stoptime = text.split("value:")[1]
    rList.append(stoptime)
    # print(rList)
    rglist = getRGlist(wlist)
    # print(rglist)
    for rl in rglist:
        t = []
        t.extend(rList)
        t.extend(rl)
        retList.append(t)
        # print(t)
    # for onecontext in retList:
    #    print(onecontext)
    return retList
    # exit(2)


def gen_excelbyfile(cdr):
    path = os.path.abspath('.')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "CDR"
    titlelist = ["序号", "话单类型", "测试号码", "IMSI", "APN", "RAT Type", "startTime", "stopTime", "RG开始时间", "RG结束时间", "上行流量",
                 "下行流量", "ServiceID", "网元名称(话单验证的网元)", "话单文件名", "CharingID"]
    x = 1
    for v in titlelist:
        sheet.cell(row=1, column=x, value=v)
        x += 1
    # y=1
    for r in range(0, len(cdr)):
        # print(key)
        sheet.cell(r + 2, 1, str(r + 1))
        for c in range(0, len(cdr[r])):
            sheet.cell(r + 2, c + 2, cdr[r][c])
            # y+=1
            # print(keyValueList)
    l_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    path = path + '\\' + 'Generated\\' + '\\计费验证\\' + l_time
    processL347.mkdir(path)
    path = path + '\\' + l_time + '.xlsx'
    wb.save(path)
    wb.close()


def writeRowInExcel(sheet, billFileName, writeList):
    py = sheet.max_row + 1
    contextList = []
    contextList = handleOneRecord(writeList)
    for onecontext in contextList:
        onecontext.append(billFileName)
        px = 2
        for value in onecontext:
            sheet.cell(row=py, column=px, value=value)
            px += 1
        py += 1
        # print(onecontext)

    '''
    for row in contextList:
        sheet.cell(row=py, column=px, value=text)
        px+= 1
    '''


# G:\LOG\log_20180207\CG29_20180207.txt
def gen_CHG(filepath, filename):
    Tflag = False
    # filepath = input("输入话单log文件")
    # filepath = "G:\LOG\log_20180126\CG29_20180126_0.txt"
    # filepath = "G:\LOG\log_20180126\CG48_20180126_0.log"
    # filepath = "G:\LOG\log_20180201\CG29_20180202_02.log"
    # filepath = "G:\LOG\log_20180201\CG29_20180201_01.log"
    chargingDict = {}
    list = []
    file = open(filepath)

    # print(file.read())
    lt = file.readlines()
    l = []
    for linestr in lt:
        # print(linestr)
        l.append(linestr.split('\n')[0])
        if linestr.find(']#') != -1 or linestr.find(']$') != -1:
            Tflag = True
    lt = []
    file.close()
    # for linestr in l:
    #   print(linestr)
    # print(len(l))
    # exit(1)

    if Tflag:
        list = getWholeChargingListByfile(l)
        for bill in list:
            recordlist = []
            f, e = bill
            billfile = getBillFileName(l[f])
            # if billfile == "cg29_pgw2_20180202012627_00061886.dat":
            # print(f,e)
            recordlist = getRecordList(l, f + 1, e + 1)
            chargingDict[billfile] = recordlist
            # print(billfile)
            # print(recordlist)
            gc.collect()
            write07Excel("话单验证", chargingDict)
    else:
        list = get_tup(l)
        cdr = []
        for tup in list:
            rcdtype = 'null'
            imsi = 'null'
            apn = 'null'
            st = 'null'
            msi = 'null'
            rat = 'null'
            rg = 'null'
            et = 'null'
            rg_et = 'null'
            rg_st = 'null'
            upvol = 'null'
            downvol = 'null'
            ne = 'null'
            sid = 'null'
            cid = 'null'
            for line in l[tup[0]:tup[1]]:
                if line.find('-->recordType(0)') != -1:
                    rcdtype = line[line.find(':') + 2:]
                if line.find('servedIMSI(3)') != -1:
                    imsi = line[line.find('x') + 1:]
                    m = ''
                    for i in range(0, len(imsi), 2):
                        m = m + imsi[i + 1] + imsi[i]
                    imsi = m[0:-1]
                if line.find('-->accessPointNameNI(7)') != -1:
                    apn = line[line.find(':') + 2:]
                if line.find('-->chargingID(5)') != -1:
                    cid = line[line.find(':') + 2:]
                if line.find('-->recordOpeningTime(13)') != -1:
                    st = line[line.find(':') + 2:]
                if line.find('-->nodeID(18)') != -1:
                    ne = line[line.find(':') + 2:]
                if line.find('-->servedMSISDN(22)') != -1:
                    msi = line[line.find('x') + 1:]
                    m = ''
                    for i in range(0, len(msi), 2):
                        m = m + msi[i + 1] + msi[i]
                    msi = m[0:-1]
                if line.find('-->rATType(30)') != -1:
                    rat = line[line.find(':') + 2:]
                if line.find('------>ratingGroup(1)') != -1:
                    rg = line[line.find(':') + 2:]
                if line.find('------>timeOfReport(14)') != -1:
                    et = line[line.find(':') + 2:]
                if line.find('------>timeOfFirstUsage(5)') != -1:
                    rg_st = line[line.find(':') + 2:]
                if line.find('------>timeOfLastUsage(6)') != -1:
                    rg_et = line[line.find(':') + 2:]
                if line.find('------>datavolumeFBCUplink(12)') != -1:
                    upvol = line[line.find(':') + 2:]
                if line.find('------>datavolumeFBCDownlink(13)') != -1:
                    downvol = line[line.find(':') + 2:]
                if line.find('------>serviceIdentifier(17)') != -1:
                    sid = line[line.find(':') + 2:]
                    rowlist = [rcdtype, msi, imsi, apn, rat, st, et, rg_st, rg_et, upvol, downvol, sid, ne, filename, cid]
                    cdr.append(rowlist)
                    rg_et = 'null'
                    rg_st = 'null'
                    upvol = 'null'
                    downvol = 'null'
                    sid = 'null'
        gen_excelbyfile(cdr)
    del l
    del lt
    del list
