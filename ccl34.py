import openpyxl


def getServiceListByList(sheet, startRow):
    layerLag_col = 1
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    layerLag = ""
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        layerLag = sheet.cell(row=rowNumber, column=layerLag_col).value
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        if protocolNumber == None and portNumberL4 != None:
            retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, "6", portNumberL4, urlL7))
            retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, "17", portNumberL4, urlL7))
        else:
            retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))

    return retList


def add_pruListtoPRUList(alllist, pruStr, configureList):
    retList = []
    for line in range(0, len(configureList) + 1):
        try:
            if pruStr in configureList[line] and "charging-rule-unit" not in configureList[line]:
                front = line
                for i in range(front, len(configureList) + 1):
                    if "exit" in configureList[i] and (
                            "policy-rule-unit" in configureList[i + 1] or "charging-rule-unit" in configureList[i + 1]):
                        end = i
                        break
        except:
            pass

    try:
        for i in range(front, end + 1):
            retList.append(configureList[i])
    except:
        pass
    if len(retList) != 0:
        alllist.append(retList)



def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    # print(sList)
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList


def judgePruL4InConfigureList(sName, clst):
    try:
        for i in len(clst):
            if 'policy-rule-unit "PRU_' + sName + '_L4"' in clst[i] and "shallow-inspection-only" in clst[i + 1]:
                return True
    except:
        pass
    return False


def judgeL4Service(lst, clist):
    pruBool = judgePruL4InConfigureList(lst[0][3], clist)
    if pruBool == True:
        return True

    for text in lst:
        if text[0] == "L4":
            return True
    return False


def handleL4ServiceList(rLst, confList):
    retList = []
    tempList = []
    for serList in rLst:
        if judgeL4Service(serList, confList) == True:
            for text in serList:
                layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = text
                layerLag = "L4"
                tempList.append(
                    (layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url))
            retList.append(tempList)
            tempList = []
        else:
            for text in serList:
                layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = text
                tempList.append(
                    (layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url))
            retList.append(tempList)
            tempList = []
    return retList


def getTheFlowNumId(serviceName):
    global serviceFlowNumListDict
    if len(serviceFlowNumListDict[serviceName]) == 0:
        if serviceName in serviceFlowNumListDict:
            if len(serviceFlowNumListDict[serviceName]) == 0:
                t = []
                serviceFlowNumListDict[serviceName].append(t)
        return 1
    else:
        if serviceFlowNumListDict[serviceName][-1][-1] + 1 > 255:
            retNum = serviceFlowNumListDict[serviceName][-1][-1] + 1 - 255
            l = []
            l.append(retNum)
            serviceFlowNumListDict[serviceName].append(l)
        else:
            retNum = serviceFlowNumListDict[serviceName][-1][-1] + 1
        return retNum


def addTheCommandtoList(lst, tup, pruLst):
    layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
    global log_list
    log_list.append("将该条目:"+str(tup)+"加入命令列表中"+"\n")
    flowId = getTheFlowNumId(tup[3])
    serviceFlowNumListDict[serviceName][-1].append(flowId)
    # 创建PRU
    lst.append('exit all' + "\n")
    lst.append("configure mobile-gateway profile policy-options " + "\n")
    global serviceDict
    if len(serviceFlowNumListDict[serviceName]) < 2:
        pruKey = "PRU_" + serviceName + "_" + layerLag
    else:
        pruKey = "PRU_" + serviceName + "_" + layerLag + "_0" + str(len(serviceFlowNumListDict[serviceName]) - 1)
    if serviceDict[pruKey] == True:
        pruStr = 'policy-rule-unit "' + pruKey + '"' + "\n"
    else:
        # pruStr = 'policy-rule-unit "' + pruKey+'" create' + "\n"
        pruStr = 'policy-rule-unit "' + pruKey + '"' + "\n"
        serviceDict[pruKey] = True


    lst.append(pruStr)
    lst.append('shallow-inspection-only' + "\n")
    lst.append('flow-description ' + str(flowId) + "\n")
    lst.append('match' + "\n")
    if layerLag == "L4":
        if protocolNumber != None:
            lst.append('protocol ' + str(protocolNumber) + "\n")
    if ipAddress!=None:
        if "/" in ipAddress:
            ip = ipAddress
        else:
            ip = ipAddress + "/32"
        lst.append('remote-ip ' + ip + "\n")
    if layerLag == "L4":
        if portNumber != None:
            if "--" in str(portNumber):
                lst.append('remote-port range ' + str(portNumber).replace("--", " ") + "\n")
            else:
                lst.append('remote-port eq ' + str(portNumber) + "\n")
    lst.append("\n")


def getTheFlowNumberList(PruList):
    retList = []
    for pruline in PruList:
        templst = []
        for text in pruline:
            if "flow-description " in text:
                templst.append(int(text.replace("\n", "").split("flow-description ")[1]))
        retList.append(templst)
    return retList


def getDeletePruStrFlowNumberByList(tup, lst):
    prustr = None
    retNumber = None

    # 有无子网掩码需要处理过
    if "/" in tup[4]:
        ip = tup[4]
    else:
        ip = tup[4] + "/32"
    del_ip_str = ip.replace(" ", "")

    # 端口范围分隔符
    portSplitStr = "--"
    if tup[6] != None:
        if portSplitStr in str(tup[6]):
            portStr = "remote-port range " + tup[6].split(portSplitStr)[0] + " " + tup[6].split(portSplitStr)[1]
        else:
            portStr = "remote-port eq " + str(tup[6])

    for i in range(0, len(lst)):
        try:

            if tup[0] == "L3":
                if "remote-ip " + del_ip_str in lst[i]:
                    retNumber = int(lst[i - 2].split("flow-description ")[1])
                    break
            elif tup[0] == "L4":
                if "protocol " + str(tup[5]) in lst[i] and "remote-ip " + del_ip_str in lst[i + 1] and portStr in lst[
                    i + 2]:
                    retNumber = int(lst[i - 2].split("flow-description ")[1])
                    break
        except:
            pass

    if retNumber != None:
        return (lst[0].replace("\n", "").split("policy-rule-unit ")[1], retNumber)

    return (None, None)


def deleteTheFlow(comList, tup, flowStrList):
    global log_list

    prustr, del_num = None, None
    for line in flowStrList:
        prustr, del_num = getDeletePruStrFlowNumberByList(tup, line)
        if prustr != None:
            log_list.append("删除该条目:" + str(tup) + "获取删除的pru名，flow id" + prustr + "," + str(del_num) + "\n")
            break

    if del_num == None:
        comList.append(str(tup) + "该删除条目在配置文件中不存在\n")
        log_list.append(str(tup) + "该删除条目在配置文件中不存在\n")
    else:
        comList.append('exit all' + "\n")
        comList.append("configure mobile-gateway profile policy-options " + "\n")
        comList.append('policy-rule-unit ' + prustr + "\n")
        comList.append("no flow-description " + str(del_num) + "\n")
        comList.append('\n')
        log_list.append(str(tup) + "删除的是:" + prustr + "下的flow " + str(del_num)+"\n")


def setTheFlowNumList(pruLst):
    if len(pruLst) > 0:
        pruKey = pruLst[0].split("policy-rule-unit ")[1].replace('\n', "")

        if pruKey not in serviceFlowNumListDict:
            serviceFlowNumListDict[pruKey] = []
        # 列表中的0 是否有0表示配置log中的flow是否已经记录过,在新增中是动态添加的
        if 0 not in serviceFlowNumListDict[pruKey]:
            serviceFlowNumListDict[pruKey].append(0)
            for pruline in pruLst:
                if "flow-description " in pruline:
                    serviceFlowNumListDict[pruKey].append(int(pruline.split("flow-description ")[1]))
        serviceFlowNumListDict[pruKey] = list(set(serviceFlowNumListDict[pruKey]))


def setPRUCRUtoServiceDict(tup, cfglst):
    global serviceDict
    serviceId = tup[2]
    serviceName = tup[3]
    if "PR_" + serviceName not in serviceDict:
        serviceDict["PR_" + serviceName] = False
    if "PRU_" + serviceName + "_" + tup[0] not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = False
    if "PRU_" + serviceName + "_" + tup[0] + "_01" not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_01"] = False
    if "PRU_" + serviceName + "_" + tup[0] + "_02" not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_02"] = False
    if "PRU_" + serviceName + "_" + tup[0] + "_03" not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_03"] = False
    if "CRU_" + serviceName not in serviceDict:
        serviceDict["CRU_" + serviceName] = False
    if serviceName == "lly_00":
        print("lly_00:++++"+"PRU_" + serviceName + "_" + tup[0],serviceDict["PRU_" + serviceName + "_" + tup[0]])
    # 判断PR是否存在
    prStr = 'policy-rule "PR_' + serviceName + '"'
    if PR_str_isExist(prStr, cfglst) == True:
        serviceDict["PR_" + serviceName] = True

    # 判断PRU是否存在
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = True
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + "_01" + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_01"] = True
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + "_02" + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_02"] = True
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + "_03" + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0] + "_03"] = True

    # 判断CRU是否存在
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    if CRU_str_isExist(cruStr, serviceId, cfglst) == True:
        serviceDict["CRU_" + serviceName] = True


def PR_str_isExist(pr_str, cfglst):
    i = 0
    for text in cfglst:
        if pr_str in text and "qci * arp * precedence" not in text:
            i += 1
    if i == 2:
        return True
    else:
        return False


def PRU_str_isExist(pru_str, cfglst):
    for text in cfglst:
        if pru_str in text and "qci * arp * precedence" not in text:
            return True
    return False


def CRU_str_isExist(cru_str, sid, cfglst):
    global commandList

    for i in range(0, len(cfglst)):
        if cru_str in cfglst[i] and "qci * arp * precedence" not in cfglst[i]:
            # print(i,cfglst[i+1])
            try:
                if cfglst[i + 1].split("rating-group ")[1].replace("\n", "") != str(sid):
                    # print(cfglst[i+1].split("rating-group ")[1].replace("\n",""))
                    print(cru_str + "该ID：" + str(sid) + "匹配不对")
                    commandList.append("注意\n")
                    commandList.append(cru_str + "该ID：" + str(sid) + "匹配不对\n")
            except:
                pass
            return True


def PRU_CRU_is_Associate(serviceName, prStr, pruStr, clst):
    prStr = 'policy-rule "' + prStr + '"'
    prustr = 'policy-rule-unit "' + pruStr + '"'
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    for text in clst:
        if prStr in text and prustr in text and cruStr in text:
            return True
    return False


def getPRUlistByConfigureList(tup, configureList):
    layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + layerLag
    exitStr = ""
    findLag = False
    retList = []
    tlist = []

    for line in range(0, len(configureList) + 1):
        try:
            if pruStr in configureList[line] and "charging-rule-unit" not in configureList[line]:
                front = line
                for i in range(front, len(configureList) + 1):
                    if "exit" in configureList[i] and (
                            "policy-rule-unit" in configureList[i + 1] or "charging-rule-unit" in configureList[i + 1]):
                        end = i
                        break
                try:
                    for i in range(front, end + 1):
                        tlist.append(configureList[i])
                    retList.append(tlist)
                    tlist = []
                    front = -1
                    end = -1
                except:
                    pass
        except:
            pass
    return retList


def PR_PRU_CRU_Process(lst, tup, cfglst):
    global log_list
    # 检测CRU是否存在，不存在则创建
    if serviceDict["CRU_" + tup[3]] == False:
        # 创建CRU
        log_list.append("创建该业务:"+tup[3]+"的CRU")
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        lst.append('charging-rule-unit "CRU_' + tup[3] + '" ' + "\n")
        lst.append('rating-group ' + str(tup[2]) + "\n")
        lst.append('service-identifier ' + str(tup[2]) + "\n")
        lst.append('reporting-level service-id' + "\n")
        lst.append('exit' + "\n")
        lst.append("\n")


    # 检测PRU是否存在，存在则关联,PRU在创建flow的时候已经创建并标记了,所以只要将现存的PRU关联下就可以了
    log_list.append("创建该业务:" + tup[3] + "的PRU")
    if len(serviceFlowNumListDict[tup[3]]) == 4:
        for i in range(4, 1, -1):
            pruStr = 'PRU_' + tup[3] + '_' + tup[0] + '_0' + str(i - 1)
            prStr = 'PR_' + tup[3] + '_0' + str(i - 1)
            id = 10000
            # print(pruStr, PRU_CRU_is_Associate(tup[3],prStr, pruStr, cfglst))
            if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
                cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                    3] + '" qci * arp * precedence ' + str(id)
                lst.append("该业务需要创建PR\n")
                lst.append(cmpstr + '\n')
                lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('exit all' + "\n")

        pruStr = 'PRU_' + tup[3] + '_' + tup[0]
        prStr = 'PR_' + tup[3]
        id = 10000
        if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
            cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                3] + '" qci * arp * precedence ' + str(id)
            lst.append("该业务需要创建PR\n")
            lst.append(cmpstr + '\n')
            lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('exit all' + "\n")
    elif len(serviceFlowNumListDict[tup[3]]) == 3:
        for i in range(3, 1, -1):
            pruStr = 'PRU_' + tup[3] + '_' + tup[0] + '_0' + str(i - 1)
            prStr = 'PR_' + tup[3] + '_0' + str(i - 1)
            id = 10000
            if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
                cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                    3] + '" qci * arp * precedence ' + str(id)
                lst.append("该业务需要创建PR\n")
                lst.append(cmpstr + '\n')
                lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('exit all' + "\n")
        pruStr = 'PRU_' + tup[3] + '_' + tup[0]
        prStr = 'PR_' + tup[3]
        id = 10000
        if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
            cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                3] + '" qci * arp * precedence ' + str(id)
            lst.append("该业务需要创建PR\n")
            lst.append(cmpstr + '\n')
            lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('exit all' + "\n")
    elif len(serviceFlowNumListDict[tup[3]]) == 2:
        for i in range(2, 1, -1):
            pruStr = 'PRU_' + tup[3] + '_' + tup[0] + '_0' + str(i - 1)
            prStr = 'PR_' + tup[3] + '_0' + str(i - 1)
            id = 10000
            if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
                cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                    3] + '" qci * arp * precedence ' + str(id)
                lst.append("该业务需要创建PR\n")
                lst.append(cmpstr + '\n')
                lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
                lst.append('policy-rule  "' + prStr + '"\n')
                lst.append('exit all' + "\n")
        pruStr = 'PRU_' + tup[3] + '_' + tup[0]
        prStr = 'PR_' + tup[3]
        id = 10000
        if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
            cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                3] + '" qci * arp * precedence ' + str(id)
            lst.append("该业务需要创建PR\n")
            lst.append(cmpstr + '\n')
            lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('exit all' + "\n")
    else:
        pruStr = 'PRU_' + tup[3] + '_' + tup[0]
        prStr = 'PR_' + tup[3]
        id = 10000
        if PRU_CRU_is_Associate(tup[3], prStr, pruStr, cfglst) == False:
            cmpstr = 'policy-rule "' + prStr + '" policy-rule-unit "' + pruStr + '" charging-rule-unit "CRU_' + tup[
                3] + '" qci * arp * precedence ' + str(id)
            lst.append("该业务需要创建PR\n")
            lst.append(cmpstr + '\n')
            lst.append('policy-rule-base  "PRB_cmnet_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('policy-rule-base  "PRB_cmwap_L3L4"' + "\n")
            lst.append('policy-rule  "' + prStr + '"\n')
            lst.append('exit all' + "\n")
    lst.append('\n')


def gen_l34(configList,path):
    # 全局变量字典用于存储业务是否已经创建（PRU,CRU）(True,False) 业务名:pru是否存在
    global serviceDict
    serviceDict = {}
    global log_list
    log_list = []
    global serviceFlowStrListDict
    serviceFlowStrListDict = {}
    global serviceFlowNumListDict
    serviceFlowNumListDict = {}

    commandList = []
    excel_path = path + "\\内容计费整理L34.xlsx"
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel["L34"]
    serviceList = []
    serviceList = getServiceListByList(sheet, 1)

    resultList = []
    resultList = arrangeTheList(serviceList)
    commandList.append('exit all' + "\n")
    commandList.append("configure mobile-gateway profile policy-options " + "\n")

    for resultlst in resultList:
        #commandList.append(resultlst[0][3] + "业务进行增删操作\n")
        setPRUCRUtoServiceDict(resultlst[0], configList)
        pruList = []
        if resultlst[0][3] not in serviceFlowStrListDict:
            pruList = getPRUlistByConfigureList(resultlst[0], configList)
            log_list.append("获取该业务："+resultlst[0][3]+"的所有PRU："+str(pruList)+"\n")
            serviceFlowStrListDict[resultlst[0][3]] = pruList
        pruFlowNumList = []
        if resultlst[0][3] not in serviceFlowNumListDict:
            pruFlowNumList = getTheFlowNumberList(pruList)
            log_list.append("获取该业务：" + resultlst[0][3] + "的所有PRU的数字列表：" + str(pruFlowNumList) + "\n")
            serviceFlowNumListDict[resultlst[0][3]] = pruFlowNumList
        for tupline in resultlst:
            if tupline[1] == "新增":
                addTheCommandtoList(commandList, tupline, serviceFlowNumListDict[tupline[3]])
            else:
                deleteTheFlow(commandList, tupline, serviceFlowStrListDict[tupline[3]])

                # for text in pruList:
        # 创建PR,CRU,PRU,关联PR
        PR_PRU_CRU_Process(commandList, resultlst[0], configList)

    fo = open(path + "\\L34.txt", "w",encoding='utf-8')
    fo.writelines(commandList)
    fo.close()
    if log_list:
        fo_log = open(path + "\\L34.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()

        fo_log = open("tmp\\L34.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()

