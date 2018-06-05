import openpyxl
import ChargingContextAai
import time


def gen_l7(configList,path):

    global max_entry_id
    global log_list
    global allEntryIdList
    global serviceDict
    serviceDict = {}
    global serviceEntryIdDict
    serviceEntryIdDict = {}
    global allEntryIdDict
    allEntryIdDict = {}
    global servicePortListDict
    servicePortListDict = {}
    global portListCommandList
    portListCommandList = []

    ccl7_cfg = open(path + '\\configureL7.log', 'r')
    ccl7_cfg_list = ccl7_cfg.readlines()
    ccl7_cfg.close()
    allEntryIdList = eval(ccl7_cfg_list[0])
    serviceDict = eval(ccl7_cfg_list[1])
    serviceEntryIdDict = eval(ccl7_cfg_list[2])
    allEntryIdDict = eval(ccl7_cfg_list[3])
    servicePortListDict = eval(ccl7_cfg_list[4])
    log_list = []
    serviceDict = {}
    serviceEntryIdDict = {}
    max_entry_id = -1
    commandList = []
    excel_path = path+"\\内容计费整理L7.xlsx"

    excel = openpyxl.load_workbook(excel_path)
    sheet = excel["L7"]
    serviceList = []
    # 该函数会根据分割符来把一条条目中包含port range的分成若干条
    serviceList = getServiceListByList(sheet, 1)

    resultList = []
    resultList = arrangeTheList(serviceList)


    resultList = arrangeTheList_2(resultList)


    log_list.append("获取所有entry id:"+str(allEntryIdList)+"\n")

    for resultlst in resultList:

        setTheServicePortListDict(resultlst[0][3],servicePortListDict,configList)
        setPRUCRUtoServiceDict(resultlst[0], configList)
        chg_app_create(resultlst[0][3], commandList)
        dnsMatchList = []
        # dnsMatchList = getNDSMatchListByConfigureLog(resultlst,configList)
        addNDSMatch2CommandList(commandList, dnsMatchList)
        addEntryIdtoserviceEntryIdDict(resultlst[0][3], configList)
        for tupline in resultlst:
            if tupline[1] == "新增":
                entryId = getTheCompatibleEntryIdByDict(tupline)
                log_list.append("获取合适的entry id："+str(entryId)+"\n")
                addTheCommandtoList_Entry(commandList, tupline, entryId)
                log_list.append("添加entry：" + str(tupline) + "\n")

            else:
                # 删除entry

                # 获取该业务的所有entry存于字典中,用来判断该业务是否删干净

                deleteEntryId = getTheDeleteEntryId(tupline, configList)
                if deleteEntryId == -1:
                    commandList.append(str(tupline)+":该条目没找到\n")
                    log_list.append(str(tupline)+":该条目没找到\n")
                else:
                    log_list.append("获取删除的的entry id：" + str(deleteEntryId) + "\n")
                    DeleteTheEntry(commandList, tupline, deleteEntryId)
                LimitTheServiceSpeed(commandList, tupline)

                # break

            PR_PRU_CRU_Process(commandList, resultlst[0], configList)
            PR_PRU_CRU_Delete(commandList, resultlst[0], configList)

        commandList.append("\n\n")

    config_stats=[]
    config_name=[]

    text_cfg = []
    text_cfg.append(str(allEntryIdList) + "\n")
    text_cfg.append(str(serviceDict) + "\n")
    text_cfg.append(str(serviceEntryIdDict) + "\n")
    text_cfg.append(str(allEntryIdDict) + "\n")
    text_cfg.append(str(servicePortListDict) + "\n")


    file = open(path + "\\configureL7.log", "w")
    file.writelines(text_cfg)
    file.close()

    portListCommandList.append("\n")
    commandList.append('exit all\n')
    fo = open(path+"\\L7.txt", "w", encoding='utf-8')
    fo.writelines(portListCommandList+commandList)
    fo.close()
    if log_list:
        fo_log = open(path + "\\L7.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()

        fo_log = open("tmp\\L7.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()
    return serviceDict


def getTheServicePortList(_service_name,_configure_list):
    retList = []
    for i in range(0,len(_configure_list)):
        if "port-list" in _configure_list[i] and _service_name in _configure_list[i] and "create" in _configure_list[i]:
            k = i
            for j in range(k,len(_configure_list)):
                if "port" in _configure_list[i]:
                    retList.append(int(_configure_list[i].replace("\n","").split("port ")[1]))
                if "exit" in _configure_list[i]:
                    break
        else:
            return None



def setTheServicePortListDict(service_name,service_port_list_dict,configure_list):

    if service_name not in service_port_list_dict:
        service_port_list_dict[service_name] = getTheServicePortList(service_name,configure_list)





def getServiceListByList(sheet, startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value
        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if portNumberL4 != None:
            if portsplit_2 in str(portNumberL4):
                portstr = portNumberL4.split(portsplit_2)[0]
                portL4List.append(portNumberL4.split(portsplit_2)[1])
            else:
                pass
            if portsplit_1 in str(portstr):
                portstr = portstr.split(portsplit_1)
                for p in portstr:
                    portL4List.append(p)
            else:
                portL4List.append(portstr)
            for p in portL4List:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
        else:
            retList.append(
                (layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    retList = list(set(retList))

    return retList

def serviceTop(result_list):
    topList = ["mgspqwdx_00", "txsp_00", "ks_00", "blbl_00", "mgtv_00", "wyyyy_00", "pptv_00", "zsyyt_01",
               "pushmail_01", "fx_01", "mgsp_00", "mgzb_00", "mgyy_00", "mgyd_00"]
    retList = []
    for i in range(0,len(topList)):
        for serviceList in result_list:
            if serviceList[0][3] == topList[i]:
                tmplist = serviceList
                retList.append(tmplist)
                result_list.remove(tmplist)
    retList = retList + result_list
    return retList



def arrangeTheList(lst):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []

    newRetList = []
    newRetList = serviceTop(retList)
    return newRetList


def arrangeTheList_2(rtList):
    new_ret_list = []
    for lst in rtList:
        tmplist = []
        for tup in lst:
            if tup[1] == "新增":
                tmplist.append(tup)
        for tup in lst:
            if tup[1] == "删除":
                tmplist.append(tup)
        new_ret_list.append(tmplist)
    return new_ret_list


def getNDSMatchListByConfigureLog():
    pass


def addNDSMatch2CommandList(comdLst, dnsMaLst):
    pass




def getTheCompatibleEntryIdByDict(tup):
    global serviceEntryIdDict
    global allEntryIdList
    global allEntryIdDict
    serviceName = tup[3]
    serviceCaseStr = "no_head"
    retId = -1

    if "cxjs" in serviceName or "cxfs" in serviceName:
        for i in range(20001, 20500):
            if i not in allEntryIdDict["caixin"]:
                retId = i
                allEntryIdDict[serviceCaseStr].append(retId)
                break
    else:
        for i in range(20501,60000):
            if i not in allEntryIdDict[serviceCaseStr]:
                retId = i
                allEntryIdDict[serviceCaseStr].append(retId)
                break

    return retId

def createThePortList(_servie_name,_port_number):
    global portListCommandList
    global servicePortListDict
    #print("创建portList",_servie_name,_port_number)
    s_p_list = []
    if "," in _port_number:
        port_list = _port_number.split(",")
        for portstr in port_list:
            if "-" in portstr:
                s_p_list.append("range "+portstr.replace("-"," "))
            else:
                s_p_list.append(portstr)
    else:
        if "-" in _port_number:
            s_p_list.append("range " + _port_number.replace("-", " "))
        else:
            s_p_list.append(_port_number)
    s_p_list.sort()
    servicePortListDict[_servie_name] = s_p_list
    portListCommandList.append('exit all' + "\n")

    portListCommandList.append('port-list "app_'+_servie_name+'" create'+"\n")
    portListCommandList.append('description "'+_servie_name+'"'+"\n")
    for portnumber in s_p_list:
        portListCommandList.append('port '+str(portnumber)+"\n")



def addPortInToPortList(_servie_name,_port_number):
    global portListCommandList
    global servicePortListDict
    #print("添加端口号进portList",_servie_name,_port_number)
    if "," in _port_number:
        port_list = _port_number.split(",")
        for portstr in port_list:
            if "-" in portstr:
                if "range "+portstr.replace("-"," ") not in servicePortListDict[_servie_name]:
                    portListCommandList.append('port ' + "range "+portstr.replace("-"," ") + "\n")
                    servicePortListDict[_servie_name].append("range "+portstr.replace("-"," "))
            else:
                if portstr not in servicePortListDict[_servie_name]:
                    #print("添加端口" + str(portstr) + "进" + _servie_name + "的portlist")
                    portListCommandList.append('port ' + str(portstr) +"\n")
                    servicePortListDict[_servie_name].append(portstr)
    else:
        if "-" in _port_number:
            if "range " + _port_number.replace("-", " ") not in servicePortListDict[_servie_name]:
                portListCommandList.append('port ' + "range " + _port_number.replace("-", " ") + "\n")
                servicePortListDict[_servie_name].append("range " + _port_number.replace("-", " "))
        else:
            if _port_number not in servicePortListDict[_servie_name]:
                #print("添加端口" + str(_port_number) + "进" + _servie_name + "的portlist")
                portListCommandList.append('port ' + str(_port_number) +"\n")
                servicePortListDict[_servie_name].append(_port_number)

def putThePortNumberInToPortList(servie_name,port_number):
    global servicePortListDict

    if servicePortListDict[servie_name] == None:
        createThePortList(servie_name,port_number)
        #print(servie_name, "的port-list is ", servicePortListDict[servie_name], port_number)
    else:
        addPortInToPortList(servie_name,port_number)
        #print("add",servie_name, "的port-list is ", servicePortListDict[servie_name], port_number)

    return '"app_'+servie_name+'"'


def addTheCommandtoList_Entry(comLst, tup, enId):
    global servicePortListDict
    global portListCommandList
    layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
    comLst.append("exit all\n")
    comLst.append("configure application-assurance group 1:1 policy\n")
    comLst.append("app-filter\n")
    comLst.append("entry " + str(enId) + " create\n")
    expression_1 = None
    expression_2 = None
    http_port = None
    if url != None:
        expression_1,expression_2,http_port = ChargingContextAai.processUrl(url)

    if expression_1 !=None:
        comLst.append('expression 1 http-host eq "' + expression_1 + '"\n')
    if expression_2 != None:
        comLst.append('expression 2 http-uri eq "' + expression_2 + '"\n')
    if ipAddress == None:
        comLst.append("server-address eq 10.0.0.172/32\n")
    else:
        if "/" not in ipAddress:
            ipAddress = ipAddress + "/32"
        comLst.append('server-address eq ' + ipAddress + '\n')
    if portNumber != None:
        portNumber = str(portNumber).replace(" ","")
        if "," in portNumber or "-" in portNumber:
            comLst.append('server-port eq port-list ' + putThePortNumberInToPortList(serviceName,portNumber) + '\n')
        else:
            comLst.append('server-port eq ' + str(portNumber) + '\n')
    if http_port != None:
        comLst.append('http-port eq ' + str(http_port) + '\n')
    comLst.append('application "APP_' + serviceName + '"\n')
    comLst.append("no shutdown\n")
    comLst.append("\n")


    # 纯7L的地址（网址应该创建dns-catch）
    global max_entry_id
    if ipAddress == None and portNumber == None and tup[7] != None:
        dns_entry_id = getTheCompatibleEntryIdByDict(tup)
        comLst.append("exit all\n")
        comLst.append("configure application-assurance group 1:1 policy\n")
        comLst.append("app-filter\n")
        comLst.append("entry " + str(dns_entry_id) + " create\n")

        if expression_1 != None:
            comLst.append('expression 1 http-host eq"'+expression_1+'"\n')
        if expression_2 != None:
            comLst.append('expression 2 http-uri eq "' + expression_2 + '"\n')

        comLst.append('server-address eq dns-ip-cache "TrustedCache"\n')
        if portNumber != None:
            if " " in str(portNumber):
                comLst.append('server-port range ' + str(portNumber) + '\n')
            else:
                comLst.append('server-port eq ' + str(portNumber) + '\n')
        if http_port != None:
            comLst.append('http-port eq ' + str(http_port) + '\n')
        comLst.append('application "APP_' + serviceName + '"\n')
        comLst.append("no shutdown\n")
        comLst.append("exit\n")
        comLst.append("\n\n")
        # 因为要添加dns-catch得有两entry所以得添加2次,这里先添加一次
    #serviceEntryIdDict[tup[3]].append(enId)


def getServiceEntryList(serviceName, cLst):
    retLst = []
    tmpLst = []
    for i in range(0, len(cLst)):
        if 'application "APP_' + serviceName + '"' in cLst[i] and "create" not in cLst[i]:
            p = i
            for j in range(p, 0, -1):
                if "entry " in cLst[j]:
                    for t in range(j, p + 2):
                        tmpLst.append(cLst[t])
                    break
            retLst.append(tmpLst)
            tmpLst = []

    return retLst


def getTheDeleteEntryId(tup, cfgLst):
    entryList = []
    expression_1 = ""
    serviceAddressStr = ""
    servicePortStr = ""
    if tup[7] != None:
        expression_str = tup[7].replace("http://", "")
        if ":" in expression_str and ":*" not in expression_str:
            expression_1 = expression_str.split(":")[0]
        else:
            if "/*" in expression_str:
                expression_str = expression_str.replace("/*", "")

            try:
                expression_1 = expression_str[0:expression_str.index("/")]
            except:
                expression_1 = expression_str

    if tup[4] != None:
        if "/" not in tup[4]:
            ipstr = tup[4] + "/32"
        else:
            ipstr = tup[4]
        serviceAddressStr = ipstr

    if tup[6] != None:
        servicePortStr = str(tup[6])

    entryList = getServiceEntryList(tup[3], cfgLst)
    delete_entry_id = -1

    for entry in entryList:
        bl = judgeTheDeleteEntry(expression_1, serviceAddressStr, servicePortStr, entry)
        if bl == True:
            delete_entry_id = int(entry[0].split("entry ")[1].split(" ")[0])
            return delete_entry_id

    return delete_entry_id


def judgeTheDeleteEntry(express, serviceAddress, servicePort, ety):
    if express == "":
        for line in ety:
            if "expression 1" in line:
                return False
    else:
        for line in ety:
            if "expression 1" in line and express not in line:
                return False

    if serviceAddress == "":
        for line in ety:
            if "server-address" in line:
                return False
    else:
        for line in ety:
            if "server-address" in line and serviceAddress not in line:
                return False

    if servicePort == "":
        for line in ety:
            if "server-port" in line:
                return False
    else:
        for line in ety:
            if "server-port" in line and servicePort not in line:
                return False

    return True


def LimitTheServiceSpeed(comLst, tup):
    pass


def DeleteTheEntry(comLst, tup, entry_id):
    global serviceEntryIdDict
    comLst.append("exit all\n")
    comLst.append("configure application-assurance group 1:1 policy\n")
    comLst.append("no entry " + str(entry_id))
    comLst.append("\n")
    try:
        serviceEntryIdDict[tup[3]].remove(entry_id)
    except:
        pass


def addEntryIdtoserviceEntryIdDict(serviceName, cfglst):
    global serviceEntryIdDict
    entryIdList = []
    if serviceName in serviceEntryIdDict:
        return ""

    for i in range(0, len(cfglst)):
        if 'application "APP_' + serviceName + '"' in cfglst[i] and "create" not in cfglst[i]:
            k = i
            for j in range(k, 0, -1):
                if 'server-address eq ip-prefix-list "app_' + serviceName in cfglst[j]:
                    break
                if "entry" in cfglst[j]:
                    entryIdList.append(int(cfglst[j].split("entry ")[1].split(" create")[0]))
                    break
    serviceEntryIdDict[serviceName] = entryIdList





def setPRUCRUtoServiceDict(tup, cfglst):
    global serviceDict
    serviceId = tup[2]
    serviceName = tup[3]
    if "PR_" + serviceName not in serviceDict:
        serviceDict["PR_" + serviceName] = False
    if "PRU_" + serviceName + "_" + tup[0] not in serviceDict:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = False
    if "CRU_" + serviceName not in serviceDict:
        serviceDict["CRU_" + serviceName] = False
    if "APP_" + serviceName not in serviceDict:
        serviceDict["APP_" + serviceName] = False
    if "CHG_" + serviceName not in serviceDict:
        serviceDict["CHG_" + serviceName] = False

    # 判断PR是否存在
    prStr = 'policy-rule "PR_' + serviceName + '"'
    if PR_str_isExist(prStr, cfglst) == True:
        serviceDict["PR_" + serviceName] = True

    # 判断PRU是否存在
    pruStr = 'policy-rule-unit "PRU_' + serviceName + '_' + tup[0] + '"'
    if PRU_str_isExist(pruStr, cfglst) == True:
        serviceDict["PRU_" + serviceName + "_" + tup[0]] = True

    # 判断CRU是否存在
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    if CRU_str_isExist(cruStr, serviceId, cfglst) == True:
        serviceDict["CRU_" + serviceName] = True

    appStr = 'application "APP_' + serviceName + '"'
    if APP_str_isExist(appStr, cfglst) == True:
        serviceDict["APP_" + serviceName] = True

    chgStr = 'application "CHG_' + serviceName + '"'
    if CHG_str_isExist(chgStr, cfglst) == True:
        serviceDict["CHG_" + serviceName] = True

def PR_str_isExist(pr_str, cfglst):
    i = 0
    for text in cfglst:
        if pr_str in text and "qci * arp * precedence" not in text:
            i += 1
    if i == 2:
        return True
    else:
        return False


def PRU_str_isExist(pru_str, cfglst):
    for text in cfglst:
        if pru_str in text and "qci * arp * precedence" not in text:
            return True
    return False


def CRU_str_isExist(cru_str, sid, cfglst):
    global commandList

    for i in range(0, len(cfglst)):
        if cru_str in cfglst[i] and "qci * arp * precedence" not in cfglst[i]:
            try:
                if cfglst[i + 1].split("rating-group ")[1].replace("\n", "") != str(sid):
                    commandList.append("注意\n")
                    commandList.append(cru_str + "该ID：" + str(sid) + "匹配不对\n")
            except:
                pass
            return True


def APP_str_isExist(app_str, cfglst):
    for text in cfglst:
        if app_str in text:
            return True
    return False


def CHG_str_isExist(chg_str, cfglst):
    for text in cfglst:
        if chg_str in text:
            return True
    return False


def PRU_CRU_is_Associate(serviceName, prStr, pruStr, clst):
    prStr = 'policy-rule "' + prStr + '"'
    prustr = 'policy-rule-unit "' + pruStr + '"'
    cruStr = 'charging-rule-unit "CRU_' + serviceName + '"'
    for text in clst:
        if prStr in text and prustr in text and cruStr in text:
            return True
    return False


def PR_PRU_CRU_Process(lst, tup, cfglst):
    global log_list
    # 检测CRU是否存在，不存在则创建
    if serviceDict["CRU_" + tup[3]] == False:
        # 创建CRU
        log_list.append("创建该业务："+tup[3]+"的CRU"+"\n")
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        lst.append('charging-rule-unit "CRU_' + tup[3] + '" ' + "\n")
        lst.append('rating-group ' + str(tup[2]) + "\n")
        lst.append('service-identifier ' + str(tup[2]) + "\n")
        lst.append('reporting-level service-id' + "\n")
        lst.append('exit' + "\n")
        lst.append("\n")
        serviceDict["CRU_" + tup[3]] = True

    # 检测PRU是否存在
    if serviceDict["PRU_" + tup[3] + '_' + tup[0]] == False:
        log_list.append("创建该业务：" + tup[3] + "的PRU" + "\n")
        lst.append('exit all' + "\n")
        lst.append("configure mobile-gateway profile policy-options " + "\n")
        pruKey = "PRU_" + tup[3] + '_' + tup[0]
        # pruStr = 'policy-rule-unit "' + pruKey + '" create' + "\n"
        pruStr = 'policy-rule-unit "' + pruKey + "\n"
        lst.append(pruStr)
        lst.append('flow-description ' + str(1) + "\n")
        lst.append('match' + "\n")
        lst.append('aa-charging-group "CHG_' + tup[3] + '"' + "\n")
        lst.append('exit' + "\n")
        lst.append('exit' + "\n")
        lst.append("\n")
        serviceDict["PRU_" + tup[3] + '_' + tup[0]] = True

    # 检测PRU,CRU是否关联到PR
    if serviceDict["PR_" + tup[3]] == False:
        log_list.append("关联该业务：" + tup[3] + "的PRU" + "\n")
        serviceDict["PR_" + tup[3]] = True
        precedenceId = 10000
        pruKey = "PRU_" + tup[3] + '_' + tup[0]
        pruStr = 'policy-rule-unit "' + pruKey + '"'
        cmpstr = 'policy-rule "PR_' + tup[3] + '" ' + pruStr + ' charging-rule-unit "CRU_' + tup[3] + '" qci * arp * precedence ' + str(precedenceId)
        lst.append(cmpstr)
        lst.append('\n')


def PRU_CRU_is_Associate(tup, clst):
    prStr = 'policy-rule "PR_' + tup[3] + '"'
    prustr = 'policy-rule-unit "PRU_' + tup[3] + '_' + tup[0] + '"'
    cruStr = 'charging-rule-unit "CRU_' + tup[3] + '"'
    for text in clst:
        if prStr in text and prustr in text and cruStr in text:
            return True
    return False


def PR_PRU_CRU_Delete(lst, tup, cfglst):
    # 检测该字典中(serviceEntryIdDict)业务的entryId 列表是否为空，若空则表示没改业务了需要删PR,PRU,CRU等
    global serviceEntryIdDict
    if len(serviceEntryIdDict[tup[3]]) == 0:
        pass
        #print(tup[3], "该业务已经删完了,需要删除相应的PRU,CRU,PR,以及删除关联")
        #lst.append(tup[3] + "该业务已经删完了,需要删除相应的PRU,CRU,PR,以及删除关联\n")


def chg_app_create(service_name, cmd_list):
    global log_list
    if serviceDict["CHG_" + service_name] == False:
        log_list.append("创建该业务:"+service_name + "的CHG" + "\n")
        cmd_list.append('exit all\n')
        cmd_list.append('configure application-assurance group 1:1 policy\n')
        cmd_list.append('charging-group "CHG_' + service_name + '" create\n')
        cmd_list.append('description "' + service_name + '_00"\n\n')

    if serviceDict["APP_" + service_name] == False:
        log_list.append("创建该业务:" + service_name + "的APP" + "\n")
        cmd_list.append('exit all\n')
        cmd_list.append('configure application-assurance group 1:1 policy\n')
        cmd_list.append('application "' + service_name + '" create\n')
        cmd_list.append('app-group "APP_GROUP_1"\n')
        cmd_list.append('charging-group "CHG_' + service_name + '"\n\n')
